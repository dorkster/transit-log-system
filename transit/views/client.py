# Copyright © 2019-2023 Justin Jacobs
#
# This file is part of the Transit Log System.
#
# The Transit Log System is free software: you can redistribute it and/or modify it under the terms
# of the GNU General Public License as published by the Free Software Foundation,
# either version 3 of the License, or (at your option) any later version.
#
# The Transit Log System is distributed in the hope that it will be useful, but WITHOUT ANY
# WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS FOR A
# PARTICULAR PURPOSE.  See the GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License along with
# The Transit Log System.  If not, see http://www.gnu.org/licenses/

import uuid
import datetime
import tempfile

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponseRedirect
from django.http import FileResponse
from django.urls import reverse
from django.core import serializers
from django.db.models import Q
from django.core.paginator import Paginator
from django.utils.http import urlencode

from transit.models import Client, Trip, Tag, TemplateTrip, SiteSettings
from transit.forms import EditClientForm, EditClientFormRestricted

from django.contrib.auth.decorators import permission_required

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

from transit.common.eventlog import *
from transit.models import LoggedEvent, LoggedEventAction, LoggedEventModel

from transit.common.util import *

@permission_required(['transit.view_client'])
def clientList(request):
    context = {
        'clients': Client.objects.all(),
    }
    return render(request, 'client/list.html', context=context)

def clientCreate(request):
    client = Client()
    return clientCreateEditCommon(request, client, is_new=True)

def clientEdit(request, id):
    client = get_object_or_404(Client, id=id)

    if request.user.has_perm('transit.change_client'):
        return clientCreateEditCommon(request, client, is_new=False)
    else:
        return clientRestrictedEdit(request, client)

@permission_required(['transit.view_client'])
def clientRestrictedEdit(request, client):
    if request.method == 'POST':
        form = EditClientFormRestricted(request.POST)

        if 'cancel' in request.POST:
            return HttpResponseRedirect(reverse('clients') + '#client_' + str(client.id))

        if form.is_valid():
            client.is_transit_policy_acknowledged = form.cleaned_data['is_transit_policy_acknowledged']
            client.save()

            log_event(request, LoggedEventAction.EDIT, LoggedEventModel.CLIENT, str(client))

            return HttpResponseRedirect(reverse('clients') + '#client_' + str(client.id))
    else:
        initial = {
            'is_transit_policy_acknowledged': client.is_transit_policy_acknowledged,
        }
        form = EditClientFormRestricted(initial=initial)

    # site_settings = SiteSettings.load()

    context = {
        'form': form,
        'client': client,
    }

    return render(request, 'client/restricted_edit.html', context)

@permission_required(['transit.change_client'])
def clientCreateEditCommon(request, client, is_new, is_dupe=False, src_trip=None, src_template_trip=None):
    if request.method == 'POST':
        form = EditClientForm(request.POST)

        if 'cancel' in request.POST:
            if src_trip != None:
                return HttpResponseRedirect(reverse('schedule', kwargs={'mode':'edit', 'year':src_trip.date.year, 'month':src_trip.date.month, 'day':src_trip.date.day}) + '#trip_' + str(src_trip.id))
            elif src_template_trip != None:
                return HttpResponseRedirect(reverse('template-trips', kwargs={'parent': src_template_trip.parent.id}) + '#trip_' + str(src_template_trip.id))
            else:
                url_hash = '' if is_new else '#client_' + str(client.id)
                return HttpResponseRedirect(reverse('clients') + url_hash)
        elif 'delete' in request.POST:
            return HttpResponseRedirect(reverse('client-delete', kwargs={'id':client.id}))

        if form.is_valid():
            prev_client = dict()
            prev_client['name'] = client.name
            prev_client['address'] = client.address
            prev_client['phone_home'] = client.phone_home
            prev_client['phone_cell'] = client.phone_cell
            prev_client['phone_alt'] = client.phone_alt
            prev_client['elderly'] = str(client.elderly)
            prev_client['ambulatory'] = str(client.ambulatory)
            prev_client['reminder_instructions'] = client.reminder_instructions
            prev_client['trip_creation_notes'] = client.trip_creation_notes

            client.name = form.cleaned_data['name']
            client.address = form.cleaned_data['address']
            client.phone_home = form.cleaned_data['phone_home']
            client.phone_cell = form.cleaned_data['phone_cell']
            client.phone_alt = form.cleaned_data['phone_alt']
            client.elderly = form.cleaned_data['elderly']
            client.ambulatory = form.cleaned_data['ambulatory']
            client.tags = form.cleaned_data['tags']
            client.staff = form.cleaned_data['staff']
            client.is_active = form.cleaned_data['is_active']
            client.is_transit_policy_acknowledged = form.cleaned_data['is_transit_policy_acknowledged']
            client.reminder_instructions = form.cleaned_data['reminder_instructions']
            client.trip_creation_notes = form.cleaned_data['trip_creation_notes']
            client.usage_style = form.cleaned_data['usage_style']

            client.save()

            if is_new:
                log_event(request, LoggedEventAction.CREATE, LoggedEventModel.CLIENT, str(client))
            else:
                log_event(request, LoggedEventAction.EDIT, LoggedEventModel.CLIENT, str(client))

            existing_clients = Client.objects.filter(name=client.name)

            try:
                update_trips = int(form.cleaned_data['update_trips'])
            except:
                update_trips = 0

            try:
                update_templates = int(form.cleaned_data['update_templates'])
            except:
                update_templates = 0

            try:
                update_method = int(form.cleaned_data['update_method'])
            except:
                update_method = 0

            if update_trips > 0 or update_templates > 0:
                update_trip_args = {
                    'update_trips': update_trips,
                    'update_trips_date': form.cleaned_data['update_trips_date'],
                    'update_templates': update_templates,
                    'update_method': update_method,
                }
                # when creating a new client, there's no previous name. So just use the current name
                if is_new:
                    prev_client['name'] = client.name
                return HttpResponseRedirect(reverse('client-update-trips', kwargs={'id': client.id}) + "?" + urlencode(prev_client) + "&" + urlencode(update_trip_args)) 
            elif existing_clients.count() > 1:
                return HttpResponseRedirect(reverse('client-fix-dupes', kwargs={'id': client.id}))
            elif src_trip != None:
                return HttpResponseRedirect(reverse('schedule', kwargs={'mode':'edit', 'year':src_trip.date.year, 'month':src_trip.date.month, 'day':src_trip.date.day}) + '#trip_' + str(src_trip.id))
            elif src_template_trip != None:
                return HttpResponseRedirect(reverse('template-trips', kwargs={'parent': src_template_trip.parent.id}) + '#trip_' + str(src_template_trip.id))
            else:
                return HttpResponseRedirect(reverse('clients') + '#client_' + str(client.id))
    else:
        initial = {
            'name': client.name,
            'address': client.address,
            'phone_home': client.phone_home,
            'phone_cell': client.phone_cell,
            'phone_alt': client.phone_alt,
            'elderly': client.elderly,
            'ambulatory': client.ambulatory,
            'tags': client.tags,
            'staff': client.staff,
            'is_active': client.is_active,
            'is_transit_policy_acknowledged': client.is_transit_policy_acknowledged,
            'reminder_instructions': client.reminder_instructions,
            'trip_creation_notes': client.trip_creation_notes,
            'usage_style': client.usage_style,
            'update_trips': False,
            'update_trips_date': datetime.date.today(),
        }
        form = EditClientForm(initial=initial)

    site_settings = SiteSettings.load()

    names = set()
    addresses = set()
    if site_settings.autocomplete_history_days > 0:
        for i in Trip.objects.filter(date__gte=(datetime.date.today() - datetime.timedelta(days=site_settings.autocomplete_history_days-1))):
            if i.name:
                names.add(str(i.name))
            if i.address:
                addresses.add(str(i.address))
            if i.destination:
                addresses.add(str(i.destination))

    context = {
        'form': form,
        'client': client,
        'is_new': is_new,
        'tags': Tag.objects.all(),
        'tags_json': serializers.serialize('json', Tag.objects.all()),
        'names': sorted(names),
        'addresses': sorted(addresses),
        'is_dupe': is_dupe,
    }

    return render(request, 'client/edit.html', context)

@permission_required(['transit.change_client'])
def clientUpdateTrips(request, id):
    client = get_object_or_404(Client, id=id)

    existing_clients = Client.objects.filter(name=client.name)

    name = request.GET.get('name')
    address = request.GET.get('address')
    phone_home = request.GET.get('phone_home')
    phone_cell = request.GET.get('phone_cell')
    phone_alt = request.GET.get('phone_alt')
    elderly = None
    ambulatory = None
    reminder_instructions = request.GET.get('reminder_instructions')

    elderly_str = request.GET.get('elderly')
    if elderly_str == 'True':
        elderly = True
    elif elderly_str == 'False':
        elderly = False

    ambulatory_str = request.GET.get('ambulatory')
    if ambulatory_str == 'True':
        ambulatory = True
    elif ambulatory_str == 'False':
        ambulatory = False

    try:
        update_trips = int(request.GET.get('update_trips'))
    except:
        # if we're on this page, we assume we're updating trips, so default to 1 (aka update all trips and templates)
        update_trips = 1

    try:
        update_templates = int(request.GET.get('update_templates'))
    except:
        # if we're on this page, we assume we're updating trips, so default to 1 (aka update all trips and templates)
        update_templates = 1

    try:
        update_trips_date = datetime.datetime.strptime(request.GET.get('update_trips_date'), '%Y-%m-%d')
    except:
        update_trips_date = None

    try:
        update_method = int(request.GET.get('update_method'))
    except:
        # default to the most broad method
        update_method = 0

    trips = []
    if update_trips > 0:
        trip_query = Trip.objects.filter(name=name, format=Trip.FORMAT_NORMAL)
        if update_trips_date:
            if update_trips == 2:
                trip_query = trip_query.filter(date__gte=update_trips_date)
            elif update_trips == 3:
                trip_query = trip_query.filter(date__lte=update_trips_date)

        for trip in trip_query:
            updated = [False for i in range(9)]

            # this is update_method == 1, but it is also used for 0
            # name
            if name != client.name:
                if trip.name == name:
                    updated[0] = True
            # address
            if address != client.address:
                if trip.address == address:
                    updated[1] = True
                if trip.destination == address:
                    updated[2] = True
            # phone numbers
            if phone_home != client.phone_home:
                if trip.phone_home == phone_home:
                    updated[3] = True
            if phone_cell != client.phone_cell:
                if trip.phone_cell == phone_cell:
                    updated[4] = True
            if phone_alt != client.phone_alt:
                if trip.phone_alt == phone_alt:
                    updated[5] = True
            # elderly/ambulatory
            if elderly != client.elderly:
                if trip.elderly == elderly:
                    updated[6] = True
            if ambulatory != client.ambulatory:
                if trip.ambulatory == ambulatory:
                    updated[7] = True
            # reminder instructions
            if reminder_instructions != client.reminder_instructions:
                if trip.reminder_instructions == reminder_instructions:
                    updated[8] = True

            if update_method == 0:
                if trip.name != client.name:
                    updated[0] = True
                if trip.phone_home != client.phone_home:
                    updated[3] = True
                if trip.phone_cell != client.phone_cell:
                    updated[4] = True
                if trip.phone_alt != client.phone_alt:
                    updated[5] = True
                if trip.elderly != client.elderly:
                    updated[6] = True
                if trip.ambulatory != client.ambulatory:
                    updated[7] = True
                if trip.reminder_instructions != client.reminder_instructions:
                    updated[8] = True

            for i in range(9):
                if updated[i]:
                    trips.append({'trip': trip, 'updated': updated})
                    break

    template_trips = []
    if update_templates > 0:
        template_trip_query = TemplateTrip.objects.filter(name=name, format=Trip.FORMAT_NORMAL)
        for trip in template_trip_query:
            updated = [False for i in range(9)]

            # this is update_method == 1, but it is also used for 0
            # name
            if name != client.name:
                if trip.name == name:
                    updated[0] = True
            # address
            if address != client.address:
                if trip.address == address:
                    updated[1] = True
                if trip.destination == address:
                    updated[2] = True
            # phone numbers
            if phone_home != client.phone_home:
                if trip.phone_home == phone_home:
                    updated[3] = True
            if phone_cell != client.phone_cell:
                if trip.phone_cell == phone_cell:
                    updated[4] = True
            if phone_alt != client.phone_alt:
                if trip.phone_alt == phone_alt:
                    updated[5] = True
            # elderly/ambulatory
            if elderly != client.elderly:
                if trip.elderly == elderly:
                    updated[6] = True
            if ambulatory != client.ambulatory:
                if trip.ambulatory == ambulatory:
                    updated[7] = True
            # reminder instructions
            if reminder_instructions != client.reminder_instructions:
                if trip.reminder_instructions == reminder_instructions:
                    updated[8] = True

            if update_method == 0:
                if trip.name != client.name:
                    updated[0] = True
                if trip.phone_home != client.phone_home:
                    updated[3] = True
                if trip.phone_cell != client.phone_cell:
                    updated[4] = True
                if trip.phone_alt != client.phone_alt:
                    updated[5] = True
                if trip.elderly != client.elderly:
                    updated[6] = True
                if trip.ambulatory != client.ambulatory:
                    updated[7] = True
                if trip.reminder_instructions != client.reminder_instructions:
                    updated[8] = True

            for i in range(9):
                if updated[i]:
                    template_trips.append({'trip': trip, 'updated': updated})
                    break

    if request.method == 'POST':
        checked_trips = request.POST.getlist('trips')
        checked_templates = request.POST.getlist('templates')

        unchecked_trips = (len(trips) != len(checked_trips))
        unchecked_templates = (len(template_trips) != len(checked_templates))

        if 'cancel' in request.POST:
            if existing_clients.count() > 1:
                return HttpResponseRedirect(reverse('client-fix-dupes', kwargs={'id': client.id}))

            return HttpResponseRedirect(reverse('clients') + '#client_' + str(client.id))
        elif 'save' in request.POST:
            for item in trips:
                if unchecked_trips and not (str(item['trip'].id) in checked_trips):
                    continue

                trip = item['trip']
                updated = item['updated']

                if updated[0]:
                    trip.name = client.name
                if updated[1]:
                    trip.address = client.address
                if updated[2]:
                    trip.destination = client.address
                if updated[3]:
                    trip.phone_home = client.phone_home
                if updated[4]:
                    trip.phone_cell = client.phone_cell
                if updated[5]:
                    trip.phone_alt = client.phone_alt
                if updated[6]:
                    trip.elderly = client.elderly
                if updated[7]:
                    trip.ambulatory = client.ambulatory
                if updated[8]:
                    trip.reminder_instructions = client.reminder_instructions

                trip.save()

            for item in template_trips:
                if unchecked_templates and not (str(item['trip'].id) in checked_templates):
                    continue

                trip = item['trip']
                updated = item['updated']

                if updated[0]:
                    trip.name = client.name
                if updated[1]:
                    trip.address = client.address
                if updated[2]:
                    trip.destination = client.address
                if updated[3]:
                    trip.phone_home = client.phone_home
                if updated[4]:
                    trip.phone_cell = client.phone_cell
                if updated[5]:
                    trip.phone_alt = client.phone_alt
                if updated[6]:
                    trip.elderly = client.elderly
                if updated[7]:
                    trip.ambulatory = client.ambulatory
                if updated[8]:
                    trip.reminder_instructions = client.reminder_instructions

                trip.save()

            trip_count = len(checked_trips) + len(checked_templates)

            if trip_count > 0:
                log_event(request, LoggedEventAction.EDIT, LoggedEventModel.CLIENT, "Updated " + str(trip_count) + " trip(s): " + str(client))

            if existing_clients.count() > 1:
                return HttpResponseRedirect(reverse('client-fix-dupes', kwargs={'id': client.id}))

            return HttpResponseRedirect(reverse('clients') + '#client_' + str(client.id))
    context = {
        'client': client,
        'trips': trips,
        'template_trips': template_trips,
        'existing_clients': existing_clients,
    }

    return render(request, 'client/update_trips.html', context)

@permission_required(['transit.change_client'])
def clientFixDupes(request, id):
    client = get_object_or_404(Client, id=id)

    existing_clients = Client.objects.filter(name=client.name)

    if request.method == 'POST':
        if 'cancel' in request.POST:
            return HttpResponseRedirect(reverse('clients') + "#client_" + str(client.id))

        for existing_client in existing_clients:
            if str(existing_client.id) in request.POST:
                clients_to_delete = existing_clients.exclude(id=existing_client.id)
                for client_to_delete in clients_to_delete:
                    log_event(request, LoggedEventAction.DELETE, LoggedEventModel.CLIENT, "Remove duplicate: " + str(client_to_delete))
                    client_to_delete.delete()
                return HttpResponseRedirect(reverse('clients') + '#client_' + str(existing_client.id))

    context = {
        'client': client,
        'existing_clients': existing_clients,
    }

    return render(request, 'client/fix_dupes.html', context)

@permission_required(['transit.delete_client'])
def clientDelete(request, id):
    client = get_object_or_404(Client, id=id)

    if request.method == 'POST':
        if 'cancel' in request.POST:
            return HttpResponseRedirect(reverse('client-edit', kwargs={'id':id}))

        log_event(request, LoggedEventAction.DELETE, LoggedEventModel.CLIENT, str(client))

        client.delete()
        return HttpResponseRedirect(reverse('clients'))

    context = {
        'model': client,
    }

    return render(request, 'model_delete.html', context)

@permission_required(['transit.change_client'])
def clientCreateFromTrip(request, trip_id):
    trip = get_object_or_404(Trip, id=trip_id)

    existing_clients = Client.objects.filter(name=trip.name)
    if existing_clients.count() > 0:
        return clientCreateEditCommon(request, existing_clients[0], is_new=False, is_dupe=True, src_trip=trip)

    client = Client()
    client.name = trip.name
    client.address = trip.address
    client.phone_home = trip.phone_home
    client.phone_cell = trip.phone_cell
    client.phone_alt = trip.phone_alt
    client.elderly = trip.elderly
    client.ambulatory = trip.ambulatory
    client.reminder_instructions = trip.reminder_instructions
    return clientCreateEditCommon(request, client, is_new=True, src_trip=trip)

@permission_required(['transit.change_client'])
def clientCreateFromTemplateTrip(request, trip_id):
    trip = get_object_or_404(TemplateTrip, id=trip_id)

    existing_clients = Client.objects.filter(name=trip.name)
    if existing_clients.count() > 0:
        return clientCreateEditCommon(request, existing_clients[0], is_new=False, is_dupe=True, src_template_trip=trip)

    client = Client()
    client.name = trip.name
    client.address = trip.address
    client.phone_home = trip.phone_home
    client.phone_cell = trip.phone_cell
    client.phone_alt = trip.phone_alt
    client.elderly = trip.elderly
    client.ambulatory = trip.ambulatory
    client.reminder_instructions = trip.reminder_instructions
    return clientCreateEditCommon(request, client, is_new=True, src_template_trip=trip)

def ajaxClientList(request):
    if not request.user.has_perm('transit.view_client'):
        return HttpResponseRedirect(reverse('login_redirect'))

    SORT_NAME = 0
    SORT_ADDRESS = 1
    SORT_PHONE_HOME = 2
    SORT_PHONE_CELL = 3
    SORT_PHONE_ALT = 4
    SORT_ELDERLY = 5
    SORT_AMBULATORY = 6
    SORT_TAGS = 7
    SORT_IS_ACTIVE = 8
    SORT_IS_TRANSIT_POLICY_ACKNOWLEDGED = 9
    SORT_USAGE_STYLE = 10
    SORT_REMINDER_INSTRUCTIONS = 11
    SORT_TRIP_CREATION_NOTES = 12

    sort_mode = request.session.get('clients_sort', SORT_NAME)
    sort_mode_dir = request.session.get('clients_sort_dir', 0)

    request_id = ''
    if request.GET['target_id'] != '':
        request_id = uuid.UUID(request.GET['target_id'])

    request_action = request.GET['target_action']
    request_data = request.GET['target_data']

    reset_current_page = False

    if request_action == 'filter_elderly':
        reset_current_page = True
        request.session['clients_elderly'] = int(request_data)
    elif request_action == 'filter_ambulatory':
        reset_current_page = True
        request.session['clients_ambulatory'] = int(request_data)
    elif request_action == 'filter_staff':
        reset_current_page = True
        request.session['clients_staff'] = int(request_data)
    elif request_action == 'filter_active':
        reset_current_page = True
        request.session['clients_active'] = int(request_data)
    elif request_action == 'filter_transit_policy':
        reset_current_page = True
        request.session['clients_transit_policy'] = int(request_data)
    elif request_action == 'filter_usage_style':
        reset_current_page = True
        request.session['clients_usage_style'] = int(request_data)
    elif request_action == 'filter_search':
        reset_current_page = True
        request.session['clients_search'] = request_data
    elif request_action == 'filter_reset':
        reset_current_page = True
        request.session['clients_elderly'] = 0
        request.session['clients_ambulatory'] = 0
        request.session['clients_staff'] = 0
        request.session['clients_active'] = 0
        request.session['clients_transit_policy'] = 0
        request.session['clients_usage_style'] = 0
        request.session['clients_search'] = ''
    elif request_action == 'toggle_extra_columns':
        request.session['clients_extra_columns'] = not request.session.get('clients_extra_columns', False)
    elif request_action == 'sort':
        new_sort_mode = int(request_data)
        if sort_mode == new_sort_mode:
            sort_mode_dir = 1 if sort_mode_dir == 0 else 0
        else:
            sort_mode_dir = 0
        sort_mode = new_sort_mode
        request.session['clients_sort'] = new_sort_mode
        request.session['clients_sort_dir'] = sort_mode_dir

    if reset_current_page:
        return render(request, 'client/ajax_reset.html', context={})

    filter_elderly = request.session.get('clients_elderly', 0)
    filter_ambulatory = request.session.get('clients_ambulatory', 0)
    filter_staff = request.session.get('clients_staff', 0)
    filter_active = request.session.get('clients_active', 0)
    filter_transit_policy = request.session.get('clients_transit_policy', 0)
    filter_usage_style = request.session.get('clients_usage_style', 0)
    filter_search = request.session.get('clients_search', '')

    clients = Client.objects.all()
    unfiltered_count = clients.count()

    if filter_elderly == 1:
        clients = clients.filter(elderly=True)
    elif filter_elderly == 2:
        clients = clients.filter(elderly=False)
    elif filter_elderly == 3:
        clients = clients.filter(elderly=None)

    if filter_ambulatory == 1:
        clients = clients.filter(ambulatory=True)
    elif filter_ambulatory == 2:
        clients = clients.filter(ambulatory=False)
    elif filter_ambulatory == 3:
        clients = clients.filter(ambulatory=None)

    if filter_staff == 1:
        clients = clients.filter(staff=True)
    elif filter_staff == 2:
        clients = clients.filter(staff=False)

    if filter_active == 1:
        clients = clients.filter(is_active=True)
    elif filter_active == 2:
        clients = clients.filter(is_active=False)

    if filter_transit_policy == 1:
        clients = clients.filter(is_transit_policy_acknowledged=True)
    elif filter_transit_policy == 2:
        clients = clients.filter(is_transit_policy_acknowledged=False)

    if filter_usage_style == 1:
        clients = clients.filter(usage_style=Client.USAGE_STYLE_REGULAR)
    elif filter_usage_style == 2:
        clients = clients.filter(usage_style=Client.USAGE_STYLE_FIELD_TRIP)

    if filter_search != '':
        clients = clients.filter(Q(name__icontains=filter_search) | Q(address__icontains=filter_search) | Q(tags__icontains=filter_search) | Q(reminder_instructions__icontains=filter_search) | Q(trip_creation_notes__icontains=filter_search))

    filtered_count = clients.count()

    if sort_mode == SORT_NAME:
        clients = clients.order_by('name')
    elif sort_mode == SORT_ADDRESS:
        clients = clients.order_by('address', 'name')
    elif sort_mode == SORT_PHONE_HOME:
        clients = clients.order_by('phone_home', 'name')
    elif sort_mode == SORT_PHONE_CELL:
        clients = clients.order_by('phone_cell', 'name')
    elif sort_mode == SORT_PHONE_ALT:
        clients = clients.order_by('phone_alt', 'name')
    elif sort_mode == SORT_ELDERLY:
        clients = clients.order_by('elderly', 'name')
    elif sort_mode == SORT_AMBULATORY:
        clients = clients.order_by('ambulatory', 'name')
    elif sort_mode == SORT_TAGS:
        clients = clients.order_by('tags', 'name')
    elif sort_mode == SORT_IS_ACTIVE:
        clients = clients.order_by('is_active', 'name')
    elif sort_mode == SORT_IS_TRANSIT_POLICY_ACKNOWLEDGED:
        clients = clients.order_by('is_transit_policy_acknowledged', 'name')
    elif sort_mode == SORT_USAGE_STYLE:
        clients = clients.order_by('usage_style', 'name')
    elif sort_mode == SORT_REMINDER_INSTRUCTIONS:
        clients = clients.order_by('reminder_instructions', 'name')
    elif sort_mode == SORT_TRIP_CREATION_NOTES:
        clients = clients.order_by('trip_creation_notes', 'name')

    if sort_mode_dir == 1:
        clients = clients.reverse()

    clients_per_page = 30
    client_pages = Paginator(list(clients), clients_per_page)
    clients_paginated = client_pages.get_page(request.GET.get('page'))
    client_page_ranges = get_paginated_ranges(page=clients_paginated, page_range=5, items_per_page=clients_per_page)

    context = {
        'clients': clients_paginated,
        'client_page_ranges': client_page_ranges,
        'filter_elderly': filter_elderly,
        'filter_ambulatory': filter_ambulatory,
        'filter_staff': filter_staff,
        'filter_active': filter_active,
        'filter_search': filter_search,
        'filter_transit_policy': filter_transit_policy,
        'filter_usage_style': filter_usage_style,
        'is_filtered': (filter_elderly > 0 or filter_ambulatory > 0 or filter_staff > 0 or filter_active > 0 or filter_search != '' or filter_transit_policy > 0 or filter_usage_style > 0),
        'filtered_count': filtered_count,
        'unfiltered_count': unfiltered_count,
        'show_extra_columns': request.session.get('clients_extra_columns', False),
        'sort_mode': sort_mode,
        'sort_mode_dir': sort_mode_dir,
    }
    return render(request, 'client/ajax_list.html', context=context)

@permission_required(['transit.view_client'])
def clientXLSX(request):
    SORT_NAME = 0
    SORT_ADDRESS = 1
    SORT_PHONE_HOME = 2
    SORT_PHONE_CELL = 3
    SORT_PHONE_ALT = 4
    SORT_ELDERLY = 5
    SORT_AMBULATORY = 6
    SORT_TAGS = 7
    SORT_IS_ACTIVE = 8
    SORT_IS_TRANSIT_POLICY_ACKNOWLEDGED = 9
    SORT_USAGE_STYLE = 10
    SORT_REMINDER_INSTRUCTIONS = 11
    SORT_TRIP_CREATION_NOTES = 12

    sort_mode = request.session.get('clients_sort', SORT_NAME)
    sort_mode_dir = request.session.get('clients_sort_dir', 0)

    filter_elderly = request.session.get('clients_elderly', 0)
    filter_ambulatory = request.session.get('clients_ambulatory', 0)
    filter_staff = request.session.get('clients_staff', 0)
    filter_active = request.session.get('clients_active', 0)
    filter_transit_policy = request.session.get('clients_transit_policy', 0)
    filter_usage_style = request.session.get('clients_usage_style', 0)
    filter_search = request.session.get('clients_search', '')

    clients = Client.objects.all()

    if filter_elderly == 1:
        clients = clients.filter(elderly=True)
    elif filter_elderly == 2:
        clients = clients.filter(elderly=False)

    if filter_ambulatory == 1:
        clients = clients.filter(ambulatory=True)
    elif filter_ambulatory == 2:
        clients = clients.filter(ambulatory=False)

    if filter_staff == 1:
        clients = clients.filter(staff=True)
    elif filter_staff == 2:
        clients = clients.filter(staff=False)

    if filter_active == 1:
        clients = clients.filter(is_active=True)
    elif filter_active == 2:
        clients = clients.filter(is_active=False)

    if filter_transit_policy == 1:
        clients = clients.filter(is_transit_policy_acknowledged=True)
    elif filter_transit_policy == 2:
        clients = clients.filter(is_transit_policy_acknowledged=False)

    if filter_usage_style == 1:
        clients = clients.filter(usage_style=Client.USAGE_STYLE_REGULAR)
    elif filter_usage_style == 2:
        clients = clients.filter(usage_style=Client.USAGE_STYLE_FIELD_TRIP)

    if filter_search != '':
        clients = clients.filter(Q(name__icontains=filter_search) | Q(address__icontains=filter_search) | Q(tags__icontains=filter_search) | Q(reminder_instructions__icontains=filter_search))

    if sort_mode == SORT_NAME:
        clients = clients.order_by('name')
    elif sort_mode == SORT_ADDRESS:
        clients = clients.order_by('address', 'name')
    elif sort_mode == SORT_PHONE_HOME:
        clients = clients.order_by('phone_home', 'name')
    elif sort_mode == SORT_PHONE_CELL:
        clients = clients.order_by('phone_cell', 'name')
    elif sort_mode == SORT_PHONE_ALT:
        clients = clients.order_by('phone_alt', 'name')
    elif sort_mode == SORT_ELDERLY:
        clients = clients.order_by('elderly', 'name')
    elif sort_mode == SORT_AMBULATORY:
        clients = clients.order_by('ambulatory', 'name')
    elif sort_mode == SORT_TAGS:
        clients = clients.order_by('tags', 'name')
    elif sort_mode == SORT_IS_ACTIVE:
        clients = clients.order_by('is_active', 'name')
    elif sort_mode == SORT_IS_TRANSIT_POLICY_ACKNOWLEDGED:
        clients = clients.order_by('is_transit_policy_acknowledged', 'name')
    elif sort_mode == SORT_USAGE_STYLE:
        clients = clients.order_by('usage_style', 'name')
    elif sort_mode == SORT_REMINDER_INSTRUCTIONS:
        clients = clients.order_by('reminder_instructions', 'name')
    elif sort_mode == SORT_TRIP_CREATION_NOTES:
        clients = clients.order_by('trip_creation_notes', 'name')

    if sort_mode_dir == 1:
        clients = clients.reverse()

    client_count = len(clients)

    temp_file = tempfile.NamedTemporaryFile()

    wb = Workbook()

    style_font_normal = Font(name='Arial', size=10)
    style_border_normal_side = Side(border_style='thin', color='FF000000')
    style_border_normal = Border(left=style_border_normal_side, right=style_border_normal_side, top=style_border_normal_side, bottom=style_border_normal_side)
    style_colwidth_normal = 13
    style_colwidth_large = 30

    style_font_header = Font(name='Arial', size=10, bold=True)
    style_alignment_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
    style_fill_header = PatternFill(fill_type='solid', fgColor='DFE0E1')
    style_rowheight_header = 25

    ws = wb.active
    ws.title = 'Clients'

    row_header = 1

    ws.row_dimensions[row_header].height = style_rowheight_header

    ws.cell(row_header, 1, 'Name')
    ws.cell(row_header, 2, 'Address')
    ws.cell(row_header, 3, 'Phone (Home)')
    ws.cell(row_header, 4, 'Phone (Cell)')
    ws.cell(row_header, 5, 'Phone (Alternate)')
    ws.cell(row_header, 6, 'Elderly?')
    ws.cell(row_header, 7, 'Ambulatory?')
    ws.cell(row_header, 8, 'Tags')
    ws.cell(row_header, 9, 'Is active?')
    ws.cell(row_header, 10, 'Transit Policy Acknowledged?')
    ws.cell(row_header, 11, 'Usage Style')
    ws.cell(row_header, 12, 'Reminder Instructions')
    ws.cell(row_header, 13, 'Trip Creation Notes')

    for i in range(0, row_header + client_count):
        row = i + 1

        # apply styles
        for col in range(1, 14):
            if col == 1 or col == 2 or col == 8 or col == 12 or col == 13:
                ws.column_dimensions[get_column_letter(col)].width = style_colwidth_large
            else:
                ws.column_dimensions[get_column_letter(col)].width = style_colwidth_normal

            ws.cell(row, col).border = style_border_normal
            if row == row_header:
                ws.cell(row, col).font = style_font_header
                ws.cell(row, col).alignment = style_alignment_header
                ws.cell(row, col).fill = style_fill_header
            else:
                ws.cell(row, col).font = style_font_normal

        if row == row_header:
            continue

        client = clients[i-1]

        ws.cell(row, 1, client.name)
        ws.cell(row, 2, client.address)
        ws.cell(row, 3, client.phone_home)
        ws.cell(row, 4, client.phone_cell)
        ws.cell(row, 5, client.phone_alt)
        ws.cell(row, 6, client.elderly)
        ws.cell(row, 7, client.ambulatory)
        ws.cell(row, 8, client.tags)
        ws.cell(row, 9, client.is_active)
        ws.cell(row, 10, client.is_transit_policy_acknowledged)
        ws.cell(row, 11, client.get_usage_style_str)
        ws.cell(row, 12, client.reminder_instructions)
        ws.cell(row, 13, client.trip_creation_notes)

    wb.save(filename=temp_file.name)

    return FileResponse(open(temp_file.name, 'rb'), filename='Transit_Clients.xlsx', as_attachment=True)
