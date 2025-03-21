# Copyright © 2019-2023 Justin Jacobs
#
# This file is part of the Transit Log System.
#
# The Transit Log System is free software: you can redistribute it and/or modify it under the terms
# of the GNU General Public License as published by the Free Software Foundation,
# either version 3 of the License, or (at your option) any later version.
#
# The Transit Log System is distributed in the hope that it will be useful, but WITHOUT ANY
# WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS FOR A
# PARTICULAR PURPOSE.  See the GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License along with
# The Transit Log System.  If not, see http://www.gnu.org/licenses/

import uuid
import datetime
import tempfile

from django.shortcuts import render, get_object_or_404
from django.http import HttpResponseRedirect
from django.http import FileResponse
from django.urls import reverse
from django.core.paginator import Paginator
from django.utils.http import urlencode

from transit.models import Destination, Trip, SiteSettings, TemplateTrip
from transit.forms import EditDestinationForm

from django.contrib.auth.decorators import permission_required

from django.db.models import Q

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

from transit.common.eventlog import *
from transit.models import LoggedEvent, LoggedEventAction, LoggedEventModel

from transit.common.util import *

@permission_required(['transit.view_destination'])
def destinationList(request):
    context = {
        'destination': Destination.objects.all(),
    }
    return render(request, 'destination/list.html', context=context)

def destinationCreate(request):
    destination = Destination()
    return destinationCreateEditCommon(request, destination, is_new=True)

def destinationEdit(request, id):
    destination = get_object_or_404(Destination, id=id)
    return destinationCreateEditCommon(request, destination, is_new=False)

@permission_required(['transit.change_destination'])
def destinationCreateEditCommon(request, destination, is_new, is_dupe=False, src_trip=None, src_template_trip=None):
    if request.method == 'POST':
        form = EditDestinationForm(request.POST)

        if 'cancel' in request.POST:
            if src_trip != None:
                return HttpResponseRedirect(reverse('schedule', kwargs={'mode':'edit', 'year':src_trip.date.year, 'month':src_trip.date.month, 'day':src_trip.date.day}) + '#trip_' + str(src_trip.id))
            elif src_template_trip != None:
                return HttpResponseRedirect(reverse('template-trips', kwargs={'parent': src_template_trip.parent.id}) + '#trip_' + str(src_template_trip.id))
            else:
                url_hash = '' if is_new else '#destination_' + str(destination.id)
                return HttpResponseRedirect(reverse('destinations') + url_hash)
        elif 'delete' in request.POST:
            return HttpResponseRedirect(reverse('destination-delete', kwargs={'id':destination.id}))

        if form.is_valid():
            prev_destination = dict()
            prev_destination['address'] = destination.address
            prev_destination['phone'] = destination.phone

            destination.address = form.cleaned_data['address']
            destination.phone = form.cleaned_data['phone']
            destination.is_active = form.cleaned_data['is_active']
            destination.save()

            if is_new:
                log_event(request, LoggedEventAction.CREATE, LoggedEventModel.DESTINATION, str(destination))
            else:
                log_event(request, LoggedEventAction.EDIT, LoggedEventModel.DESTINATION, str(destination))

            existing_destinations = Destination.objects.filter(address=destination.address)

            try:
                update_trips = int(form.cleaned_data['update_trips'])
            except:
                update_trips = 0

            try:
                update_templates = int(form.cleaned_data['update_templates'])
            except:
                update_templates = 0

            if update_trips > 0 or update_templates > 0:
                update_trip_args = {
                    'update_trips': update_trips,
                    'update_trips_date': form.cleaned_data['update_trips_date'],
                    'update_templates': update_templates,
                }
                # when creating a new destination, there's no previous address. So just use the current address
                if is_new:
                    prev_destination['address'] = destination.address
                return HttpResponseRedirect(reverse('destination-update-trips', kwargs={'id': destination.id}) + "?" + urlencode(prev_destination) + "&" + urlencode(update_trip_args))
            elif existing_destinations.count() > 1:
                return HttpResponseRedirect(reverse('destination-fix-dupes', kwargs={'id': destination.id}))
            elif src_trip != None:
                return HttpResponseRedirect(reverse('schedule', kwargs={'mode':'edit', 'year':src_trip.date.year, 'month':src_trip.date.month, 'day':src_trip.date.day}) + '#trip_' + str(src_trip.id))
            elif src_template_trip != None:
                return HttpResponseRedirect(reverse('template-trips', kwargs={'parent': src_template_trip.parent.id}) + '#trip_' + str(src_template_trip.id))
            else:
                return HttpResponseRedirect(reverse('destinations') + '#destination_' + str(destination.id))
    else:
        initial = {
            'address': destination.address,
            'phone': destination.phone,
            'is_active': destination.is_active,
            'update_trips': False,
            'update_trips_date': datetime.date.today(),
        }
        form = EditDestinationForm(initial=initial)

    site_settings = SiteSettings.load()

    addresses = set()
    if site_settings.autocomplete_history_days > 0:
        for i in Trip.objects.filter(date__gte=(datetime.date.today() - datetime.timedelta(days=site_settings.autocomplete_history_days-1))):
            if i.address:
                addresses.add(str(i.address))
            if i.destination:
                addresses.add(str(i.destination))

    context = {
        'form': form,
        'destination': destination,
        'is_new': is_new,
        'addresses': sorted(addresses),
        'is_dupe': is_dupe,
    }

    return render(request, 'destination/edit.html', context)

@permission_required(['transit.change_destination'])
def destinationUpdateTrips(request, id):
    destination = get_object_or_404(Destination, id=id)

    existing_destinations = Destination.objects.filter(address=destination.address)

    address = request.GET.get('address')
    phone = request.GET.get('phone')

    try:
        update_trips = int(request.GET.get('update_trips'))
    except:
        # if we're on this page, we assume we're updating trips, so default to 1 (aka update all trips and templates)
        update_trips = 1

    try:
        update_templates = int(request.GET.get('update_templates'))
    except:
        # if we're on this page, we assume we're updating trips, so default to 1 (aka update all trips and templates)
        update_templates = 1

    try:
        update_trips_date = datetime.datetime.strptime(request.GET.get('update_trips_date'), '%Y-%m-%d')
    except:
        update_trips_date = None

    trips = []
    if update_trips > 0:
        trip_query = Trip.objects.filter(Q(address=address) | Q(destination=address))
        if update_trips_date:
            if update_trips == 2:
                trip_query = trip_query.filter(date__gte=update_trips_date)
            elif update_trips == 3:
                trip_query = trip_query.filter(date__lte=update_trips_date)

        for trip in trip_query:
            updated = [False for i in range(4)]

            # address
            if address != destination.address:
                if trip.address == address:
                    updated[0] = True
                if trip.destination == address:
                    updated[2] = True
            # phone numbers
            if phone != destination.phone:
                if trip.phone_address == phone and trip.address == address:
                    updated[1] = True
                if trip.phone_destination == phone and trip.destination == address:
                    updated[3] = True

            for i in range(4):
                if updated[i]:
                    trips.append({'trip': trip, 'updated': updated})
                    break

    template_trips = []
    if update_templates > 0:
        template_trip_query = TemplateTrip.objects.filter(Q(address=address) | Q(destination=address))
        if update_trips_date:
            if update_trips == 2:
                template_trip_query = template_trip_query.filter(date__gte=update_trips_date)
            elif update_trips == 3:
                template_trip_query = template_trip_query.filter(date__lte=update_trips_date)

        for trip in template_trip_query:
            updated = [False for i in range(4)]

            # address
            if address != destination.address:
                if trip.address == address:
                    updated[0] = True
                if trip.destination == address:
                    updated[2] = True
            # phone numbers
            if phone != destination.phone:
                if trip.phone_address == phone and trip.address == address:
                    updated[1] = True
                if trip.phone_destination == phone and trip.destination == address:
                    updated[3] = True

            for i in range(4):
                if updated[i]:
                    template_trips.append({'trip': trip, 'updated': updated})
                    break

    if request.method == 'POST':
        checked_trips = request.POST.getlist('trips')
        checked_templates = request.POST.getlist('templates')

        unchecked_trips = (len(trips) != len(checked_trips))
        unchecked_templates = (len(template_trips) != len(checked_templates))

        if 'cancel' in request.POST:
            if existing_destinations.count() > 1:
                return HttpResponseRedirect(reverse('destination-fix-dupes', kwargs={'id': destination.id}))

            return HttpResponseRedirect(reverse('destinations') + '#destination_' + str(destination.id))
        elif 'save' in request.POST:
            for item in trips:
                if unchecked_trips and not (str(item['trip'].id) in checked_trips):
                    continue

                trip = item['trip']
                updated = item['updated']

                if updated[0]:
                    trip.address = destination.address
                if updated[1]:
                    trip.phone_address = destination.phone
                if updated[2]:
                    trip.destination = destination.address
                if updated[3]:
                    trip.phone_destination = destination.phone

                trip.save()

            for item in template_trips:
                if unchecked_templates and not (str(item['trip'].id) in checked_templates):
                    continue

                trip = item['trip']
                updated = item['updated']

                if updated[0]:
                    trip.address = destination.address
                if updated[1]:
                    trip.phone_address = destination.phone
                if updated[2]:
                    trip.destination = destination.address
                if updated[3]:
                    trip.phone_destination = destination.phone

                trip.save()

            trip_count = len(checked_trips) + len(checked_templates)

            if trip_count > 0:
                log_event(request, LoggedEventAction.EDIT, LoggedEventModel.DESTINATION, "Updated " + str(trip_count) + " trip(s): " + str(destination))

            if existing_destinations.count() > 1:
                return HttpResponseRedirect(reverse('destination-fix-dupes', kwargs={'id': destination.id}))

            return HttpResponseRedirect(reverse('destinations') + '#destination_' + str(destination.id))
    context = {
        'destination': destination,
        'trips': trips,
        'template_trips': template_trips,
        'existing_destinations': existing_destinations,
    }

    return render(request, 'destination/update_trips.html', context)

@permission_required(['transit.change_destination'])
def destinationFixDupes(request, id):
    destination = get_object_or_404(Destination, id=id)

    existing_destinations = Destination.objects.filter(address=destination.address)

    if request.method == 'POST':
        if 'cancel' in request.POST:
            return HttpResponseRedirect(reverse('destinations') + "#destination_" + str(destination.id))

        for existing_destination in existing_destinations:
            if str(existing_destination.id) in request.POST:
                destinations_to_delete = existing_destinations.exclude(id=existing_destination.id)
                for destination_to_delete in destinations_to_delete:
                    log_event(request, LoggedEventAction.DELETE, LoggedEventModel.DESTINATION, "Remove duplicate: " + str(destination_to_delete))
                    destination_to_delete.delete()
                return HttpResponseRedirect(reverse('destinations') + '#destination_' + str(existing_destination.id))

    context = {
        'destination': destination,
        'existing_destinations': existing_destinations,
    }

    return render(request, 'destination/fix_dupes.html', context)

@permission_required(['transit.delete_destination'])
def destinationDelete(request, id):
    destination = get_object_or_404(Destination, id=id)

    if request.method == 'POST':
        if 'cancel' in request.POST:
            return HttpResponseRedirect(reverse('destination-edit', kwargs={'id':id}))

        log_event(request, LoggedEventAction.DELETE, LoggedEventModel.DESTINATION, str(destination))

        destination.delete()
        return HttpResponseRedirect(reverse('destinations'))

    context = {
        'model': destination,
    }

    return render(request, 'model_delete.html', context)

@permission_required(['transit.change_destination'])
def destinationCreateFromTrip(request, trip_id, use_address):
    trip = get_object_or_404(Trip, id=trip_id)

    if use_address == 1:
        target_address = trip.address
        target_phone = trip.phone_address
    else:
        # 'destination' is the default
        target_address = trip.destination
        target_phone = trip.phone_destination

    existing_destinations = Destination.objects.filter(address=target_address)
    if existing_destinations.count() > 0:
        return destinationCreateEditCommon(request, existing_destinations[0], is_new=False, is_dupe=True, src_trip=trip)

    destination = Destination()
    destination.address = target_address
    destination.phone = target_phone
    return destinationCreateEditCommon(request, destination, is_new=True, src_trip=trip)

@permission_required(['transit.change_destination'])
def destinationCreateFromTemplateTrip(request, trip_id, use_address):
    trip = get_object_or_404(TemplateTrip, id=trip_id)

    if use_address == 1:
        target_address = trip.address
        target_phone = trip.phone_address
    else:
        # 'destination' is the default
        target_address = trip.destination
        target_phone = trip.phone_destination

    existing_destinations = Destination.objects.filter(address=target_address)
    if existing_destinations.count() > 0:
        return destinationCreateEditCommon(request, existing_destinations[0], is_new=False, is_dupe=True, src_template_trip=trip)

    destination = Destination()
    destination.address = target_address
    destination.phone = target_phone
    return destinationCreateEditCommon(request, destination, is_new=True, src_template_trip=trip)

def ajaxDestinationList(request):
    if not request.user.has_perm('transit.view_destination'):
        return HttpResponseRedirect(reverse('login_redirect'))

    SORT_ADDRESS = 0
    SORT_PHONE = 1
    SORT_IS_ACTIVE = 2

    sort_mode = request.session.get('destinations_sort', SORT_ADDRESS)
    sort_mode_dir = request.session.get('destinations_sort_dir', 0)

    request_id = ''
    if request.GET['target_id'] != '':
        request_id = uuid.UUID(request.GET['target_id'])

    request_action = request.GET['target_action']
    request_data = request.GET['target_data']

    reset_current_page = False

    if request_action == 'filter_active':
        reset_current_page = True
        request.session['destinations_active'] = int(request_data)
    elif request_action == 'filter_search':
        reset_current_page = True
        request.session['destinations_search'] = request_data
    elif request_action == 'filter_reset':
        reset_current_page = True
        request.session['destinations_active'] = 0
        request.session['destinations_search'] = ''
    elif request_action == 'sort':
        new_sort_mode = int(request_data)
        if sort_mode == new_sort_mode:
            sort_mode_dir = 1 if sort_mode_dir == 0 else 0
        else:
            sort_mode_dir = 0
        sort_mode = new_sort_mode
        request.session['destinations_sort'] = new_sort_mode
        request.session['destinations_sort_dir'] = sort_mode_dir

    if reset_current_page:
        return render(request, 'destination/ajax_reset.html', context={})

    filter_active = request.session.get('destinations_active', 0)
    filter_search = request.session.get('destinations_search', '')

    destinations = Destination.objects.all()
    unfiltered_count = destinations.count()

    if filter_active == 1:
        destinations = destinations.filter(is_active=True)
    if filter_active == 2:
        destinations = destinations.filter(is_active=False)

    if filter_search != '':
        destinations = destinations.filter(address__icontains=filter_search)

    filtered_count = destinations.count()

    if sort_mode == SORT_ADDRESS:
        destinations = destinations.order_by('address')
    elif sort_mode == SORT_PHONE:
        destinations = destinations.order_by('phone', 'address')
    elif sort_mode == SORT_IS_ACTIVE:
        destinations = destinations.order_by('is_active', 'address')

    if sort_mode_dir == 1:
        destinations = destinations.reverse()

    destinations_per_page = 30
    destination_pages = Paginator(list(destinations), destinations_per_page)
    destinations_paginated = destination_pages.get_page(request.GET.get('page'))
    destination_page_ranges = get_paginated_ranges(page=destinations_paginated, page_range=5, items_per_page=destinations_per_page)

    context = {
        'destinations': destinations_paginated,
        'destination_page_ranges': destination_page_ranges,
        'filter_active': filter_active,
        'filter_search': filter_search,
        'is_filtered': (filter_active > 0 or filter_search != ''),
        'filtered_count': filtered_count,
        'unfiltered_count': unfiltered_count,
        'sort_mode': sort_mode,
        'sort_mode_dir': sort_mode_dir,
    }
    return render(request, 'destination/ajax_list.html', context=context)

@permission_required(['transit.view_destination'])
def destinationXLSX(request):
    SORT_ADDRESS = 0
    SORT_PHONE = 1
    SORT_IS_ACTIVE = 2

    sort_mode = request.session.get('destinations_sort', SORT_ADDRESS)
    sort_mode_dir = request.session.get('destinations_sort_dir', 0)

    filter_active = request.session.get('destinations_active', 0)
    filter_search = request.session.get('destinations_search', '')

    destinations = Destination.objects.all()

    if filter_active == 1:
        destinations = destinations.filter(is_active=True)
    elif filter_active == 2:
        destinations = destinations.filter(is_active=False)

    if filter_search != '':
        destinations = destinations.filter(address__icontains=filter_search)

    if sort_mode == SORT_ADDRESS:
        destinations = destinations.order_by('address')
    elif sort_mode == SORT_PHONE:
        destinations = destinations.order_by('phone', 'address')
    elif sort_mode == SORT_IS_ACTIVE:
        destinations = destinations.order_by('is_active', 'address')

    if sort_mode_dir == 1:
        destinations = destinations.reverse()

    destination_count = len(destinations)

    temp_file = tempfile.NamedTemporaryFile()

    wb = Workbook()

    style_font_normal = Font(name='Arial', size=10)
    style_border_normal_side = Side(border_style='thin', color='FF000000')
    style_border_normal = Border(left=style_border_normal_side, right=style_border_normal_side, top=style_border_normal_side, bottom=style_border_normal_side)
    style_colwidth_normal = 13
    style_colwidth_large = 30
    style_colwidth_xlarge = 120

    style_font_header = Font(name='Arial', size=10, bold=True)
    style_alignment_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
    style_fill_header = PatternFill(fill_type='solid', fgColor='DFE0E1')
    style_rowheight_header = 25

    ws = wb.active
    ws.title = 'Destinations'

    row_header = 1

    ws.row_dimensions[row_header].height = style_rowheight_header

    ws.cell(row_header, 1, 'Address')
    ws.cell(row_header, 2, 'Phone')
    ws.cell(row_header, 3, 'Is active?')

    for i in range(0, row_header + destination_count):
        row = i + 1

        # apply styles
        for col in range(1, 4):
            if col == 1:
                ws.column_dimensions[get_column_letter(col)].width = style_colwidth_xlarge
            elif col == 1:
                ws.column_dimensions[get_column_letter(col)].width = style_colwidth_large
            else:
                ws.column_dimensions[get_column_letter(col)].width = style_colwidth_normal

            ws.cell(row, col).border = style_border_normal
            if row == row_header:
                ws.cell(row, col).font = style_font_header
                ws.cell(row, col).alignment = style_alignment_header
                ws.cell(row, col).fill = style_fill_header
            else:
                ws.cell(row, col).font = style_font_normal

        if row == row_header:
            continue

        destination = destinations[i-1]

        ws.cell(row, 1, destination.address)
        ws.cell(row, 2, destination.phone)
        ws.cell(row, 3, destination.is_active)

    wb.save(filename=temp_file.name)

    return FileResponse(open(temp_file.name, 'rb'), filename='Transit_Destinations.xlsx', as_attachment=True)
