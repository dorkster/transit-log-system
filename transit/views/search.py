# Copyright © 2019-2023 Justin Jacobs
#
# This file is part of the Transit Log System.
#
# The Transit Log System is free software: you can redistribute it and/or modify it under the terms
# of the GNU General Public License as published by the Free Software Foundation,
# either version 3 of the License, or (at your option) any later version.
#
# The Transit Log System is distributed in the hope that it will be useful, but WITHOUT ANY
# WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS FOR A
# PARTICULAR PURPOSE.  See the GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License along with
# The Transit Log System.  If not, see http://www.gnu.org/licenses/

import datetime, uuid
import tempfile

from django.shortcuts import render
from django.http import HttpResponseRedirect
from django.http import FileResponse
from django.urls import reverse
from django.core.paginator import Paginator
from django.utils.http import urlencode
from django.db.models import Q

from transit.models import Trip, Driver, Vehicle, TripType, Volunteer
from transit.forms import SearchTripsForm

from django.contrib.auth.decorators import permission_required

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

from transit.common.util import *

@permission_required(['transit.view_trip'])
def searchGetTrips(request):
    name = request.GET.get('name')
    address = request.GET.get('address')
    destination = request.GET.get('destination')
    driver = request.GET.get('driver')
    vehicle = request.GET.get('vehicle')
    start_year = request.GET.get('start_date_year')
    start_month = request.GET.get('start_date_month')
    start_day = request.GET.get('start_date_day')
    end_year = request.GET.get('end_date_year')
    end_month = request.GET.get('end_date_month')
    end_day = request.GET.get('end_date_day')
    notes = request.GET.get('notes')
    reminder_instructions = request.GET.get('reminder_instructions')
    elderly = request.GET.get('elderly')
    ambulatory = request.GET.get('ambulatory')
    trip_type = request.GET.get('trip_type')
    tags = request.GET.get('tags')
    status = request.GET.get('status')
    passenger = request.GET.get('passenger')
    volunteer = request.GET.get('volunteer')
    sort_mode = request.GET.get('sort_mode')
    result_type = request.GET.get('result_type')
    pick_up_time = request.GET.get('pick_up_time')
    appointment_time = request.GET.get('appointment_time')
    completed_log = request.GET.get('completed_log')
    fare = request.GET.get('fare')
    money_collected = request.GET.get('money_collected')
    column_layout = request.GET.get('column_layout')

    trips = Trip.objects.all().select_related('driver', 'vehicle', 'trip_type', 'volunteer')

    searched = False
    wildcard = '*'

    if name:
        name = name.strip()
        searched = True
        trips = trips.filter(format=Trip.FORMAT_NORMAL)
        if name == wildcard:
            trips = trips.exclude(name='')
        else:
            trips = trips.filter(name__icontains=name)

    if address:
        address = address.strip()
        searched = True
        trips = trips.filter(format=Trip.FORMAT_NORMAL)
        if address == wildcard:
            trips = trips.exclude(address='')
        else:
            trips = trips.filter(address__icontains=address)

    if destination:
        destination = destination.strip()
        searched = True
        trips = trips.filter(format=Trip.FORMAT_NORMAL)
        if destination == wildcard:
            trips = trips.exclude(destination='')
        else:
            trips = trips.filter(destination__icontains=destination)

    if driver:
        try:
            driver_obj = Driver.objects.get(id=uuid.UUID(driver))
        except:
            driver_obj = None

        if driver_obj:
            searched = True
            trips = trips.filter(driver=driver_obj)

    if vehicle:
        try:
            vehicle_obj = Vehicle.objects.get(id=uuid.UUID(vehicle))
        except:
            vehicle_obj = None

        if vehicle_obj:
            searched = True
            trips = trips.filter(format=Trip.FORMAT_NORMAL).filter(vehicle=vehicle_obj)

    if volunteer:
        try:
            volunteer_obj = Volunteer.objects.get(id=uuid.UUID(volunteer))
        except:
            volunteer_obj = None

        if volunteer_obj:
            searched = True
            trips = trips.filter(format=Trip.FORMAT_NORMAL).filter(volunteer=volunteer_obj)

    today = datetime.datetime.today()

    start_date = None
    if start_year:
        start_date = datetime.date(year=int(start_year), month=1, day=1)
        if start_month:
            start_date = start_date.replace(month=int(start_month))
        if start_day:
            day = int(start_day)
            if day <= 28:
                start_date = start_date.replace(day=int(start_day))
            else:
                for i in range(28, day+1):
                    try:
                        start_date = start_date.replace(day=i)
                    except:
                        pass
    elif start_month or start_day:
        start_date = datetime.date(year=today.year, month=1, day=1)
        if start_month:
            start_date = start_date.replace(month=int(start_month))
        if start_day:
            day = int(start_day)
            if day <= 28:
                start_date = start_date.replace(day=int(start_day))
            else:
                for i in range(28, day+1):
                    try:
                        start_date = start_date.replace(day=i)
                    except:
                        pass

    if start_date:
        searched = True
        trips = trips.filter(date__gte=start_date)

    end_date = None
    if end_year:
        end_date = datetime.date(year=int(end_year), month=1, day=1)
        if end_month:
            end_date = end_date.replace(month=int(end_month))
        if end_day:
            day = int(end_day)
            if day <= 28:
                end_date = end_date.replace(day=int(end_day))
            else:
                for i in range(28, day+1):
                    try:
                        end_date = end_date.replace(day=i)
                    except:
                        pass
    elif end_month or end_day:
        end_date = datetime.date(year=today.year, month=1, day=1)
        if end_month:
            end_date = end_date.replace(month=int(end_month))
        if end_day:
            day = int(end_day)
            if day <= 28:
                end_date = end_date.replace(day=int(end_day))
            else:
                for i in range(28, day+1):
                    try:
                        end_date = end_date.replace(day=i)
                    except:
                        pass

    if end_date:
        searched = True
        trips = trips.filter(date__lte=end_date)

    if notes:
        notes = notes.strip()
        searched = True
        if notes == wildcard:
            trips = trips.exclude(note='')
        else:
            trips = trips.filter(note__icontains=notes)

    if reminder_instructions:
        reminder_instructions = reminder_instructions.strip()
        searched = True
        trips = trips.filter(format=Trip.FORMAT_NORMAL)
        if reminder_instructions == wildcard:
            trips = trips.exclude(reminder_instructions='')
        else:
            trips = trips.filter(reminder_instructions__icontains=reminder_instructions)

    if elderly:
        searched = True
        if elderly == '0':
            trips = trips.filter(format=Trip.FORMAT_NORMAL).filter(elderly=None)
        elif elderly == '1':
            trips = trips.filter(format=Trip.FORMAT_NORMAL).filter(elderly=True)
        elif elderly == '2':
            trips = trips.filter(format=Trip.FORMAT_NORMAL).filter(elderly=False)

    if ambulatory:
        searched = True
        if ambulatory == '0':
            trips = trips.filter(format=Trip.FORMAT_NORMAL).filter(ambulatory=None)
        elif ambulatory == '1':
            trips = trips.filter(format=Trip.FORMAT_NORMAL).filter(ambulatory=True)
        elif ambulatory == '2':
            trips = trips.filter(format=Trip.FORMAT_NORMAL).filter(ambulatory=False)

    if trip_type:
        try:
            trip_type_obj = TripType.objects.get(id=uuid.UUID(trip_type))
        except:
            trip_type_obj = None

        if trip_type_obj:
            searched = True
            trips = trips.filter(format=Trip.FORMAT_NORMAL).filter(trip_type=trip_type_obj)

    if tags:
        searched = True
        tags = tags.strip()
        trips = trips.filter(format=Trip.FORMAT_NORMAL)
        if tags == wildcard:
            trips = trips.exclude(tags='')
        else:
            trips = trips.filter(tags__icontains=tags)

    if status:
        searched = True
        if status == '0':
            trips = trips.filter(status=Trip.STATUS_NORMAL)
        elif status == '1':
            trips = trips.filter(status=Trip.STATUS_CANCELED)
        elif status == '2':
            trips = trips.filter(status=Trip.STATUS_NO_SHOW)

    if passenger:
        searched = True
        if passenger == '1':
            trips = trips.filter(passenger=True)
        elif passenger == '2':
            trips = trips.filter(passenger=False, format=Trip.FORMAT_NORMAL)

    if pick_up_time:
        searched = True
        if pick_up_time == '1':
            trips = trips.exclude(pick_up_time='')
        elif pick_up_time == '2':
            trips = trips.filter(pick_up_time='')

    if appointment_time:
        searched = True
        if appointment_time == '1':
            trips = trips.exclude(appointment_time='')
        elif appointment_time == '2':
            trips = trips.filter(appointment_time='')

    if completed_log:
        searched = True
        if completed_log == '1':
            trips = trips.exclude(Q(start_miles='') | Q(start_time='') | Q(end_miles='') | Q(end_time=''))
        elif completed_log == '2':
            trips = trips.filter(Q(start_miles='') | Q(start_time='') | Q(end_miles='') | Q(end_time=''))

    if fare:
        searched = True
        if fare == '1':
            trips = trips.filter(fare__gt=0)
        elif fare == '2':
            trips = trips.filter(fare=0)

    if money_collected:
        searched = True
        if money_collected == '1':
            trips = trips.filter(Q(collected_cash__gt=0) | Q(collected_check__gt=0))
        elif money_collected == '2':
            trips = trips.filter(collected_cash=0, collected_check=0)

    if sort_mode:
        if sort_mode == '0':
            trips = trips.order_by('-date', '-sort_index')
        elif sort_mode == '1':
            trips = trips.order_by('date', 'sort_index')
    else:
        trips = trips.order_by('-date', '-sort_index')

    if result_type:
        if result_type == '0':
            trips = trips.filter(format=Trip.FORMAT_NORMAL)
        elif result_type == '1':
            trips = trips.filter(format=Trip.FORMAT_ACTIVITY)

    try:
        column_layout_int = int(column_layout)
    except:
        column_layout_int = 0
    return (searched, trips, column_layout_int)

@permission_required(['transit.view_trip'])
def search(request):
    results = searchGetTrips(request)
    searched = results[0]
    trips = results[1]

    results_per_page = 25
    try:
        results_per_page = int(request.GET.get('results_per_page'))
        if results_per_page == 0:
            results_per_page = 25
    except:
        pass

    result_pages = Paginator(trips, results_per_page)
    results_paginated = result_pages.get_page(request.GET.get('page'))

    page_ranges = get_paginated_ranges(page=results_paginated, page_range=5, items_per_page=results_per_page)

    try:
        column_layout = int(request.GET.get('column_layout'))
    except:
        column_layout = 0

    form = SearchTripsForm(request.GET)
    context = {
        'form': form,
        'searched': searched,
        'results': results_paginated if searched else Trip.objects.none(),
        'page_ranges': page_ranges,
        'query_string': request.GET.urlencode(),
        'Trip': Trip,
        'column_layout': column_layout,
    }
    return render(request, 'search.html', context=context)

@permission_required(['transit.view_trip'])
def searchExportXLSX(request):
    results = searchGetTrips(request)
    searched = results[0]
    trips = results[1]
    column_layout = results[2]

    temp_file = tempfile.NamedTemporaryFile()

    wb = Workbook()

    style_font_normal = Font(name='Arial', size=10)
    style_border_normal_side = Side(border_style='thin', color='FF000000')
    style_border_normal = Border(left=style_border_normal_side, right=style_border_normal_side, top=style_border_normal_side, bottom=style_border_normal_side)
    style_colwidth_normal = 13
    style_colwidth_large = style_colwidth_normal * 2
    style_colwidth_xlarge = style_colwidth_normal * 3

    style_font_header = Font(name='Arial', size=10, bold=True)
    style_alignment_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
    style_fill_header = PatternFill(fill_type='solid', fgColor='DFE0E1')
    style_rowheight_header = 25

    style_alignment_date = Alignment(horizontal='left')

    style_alignment_wrapped = Alignment(wrap_text = True, vertical = 'top')

    ws_results = wb.active
    ws_results.title = 'Search Results'

    row_header = 1
    trip_count = len(trips)

    ws_results.row_dimensions[row_header].height = style_rowheight_header

    # "Show All" columns
    COL_DATE = 1
    COL_PICKUP = 2
    COL_APPOINTMENT = 3
    COL_NAME = 4
    COL_ADDRESS = 5
    COL_PHONE = 6
    COL_DESTINATION = 7
    COL_DRIVER = 8
    COL_VEHICLE = 9
    COL_START_MILES = 10
    COL_START_TIME = 11
    COL_END_MILES = 12
    COL_END_TIME = 13
    COL_NOTES = 14
    # NOTE Reminder instructions?
    COL_TRIPTYPE = 15
    COL_PASSENGER = 16
    COL_FARE = 17
    COL_MONEY = 18
    COL_ELDERLY = 19
    COL_AMBULATORY = 20
    COL_VOLUNTEER = 21
    COL_DROPOFF_TIME = 0

    if column_layout >= 1:
        # "Standard" Columns
        COL_VOLUNTEER = 15
        COL_TRIPTYPE = 0
        COL_PASSENGER = 0
        COL_FARE = 0
        COL_MONEY = 0
        COL_ELDERLY = 0
        COL_AMBULATORY = 0
    if column_layout >= 2:
        # "Volunteer Report" Columns
        COL_NOTES = 8
        COL_VOLUNTEER = 9
        COL_DRIVER = 0
        COL_VEHICLE = 0
        COL_START_MILES = 0
        COL_START_TIME = 0
        COL_END_MILES = 0
        COL_END_TIME = 0
    if column_layout == 3:
        # "Drop-off Time Report" Columns
        COL_NAME = 2
        COL_DESTINATION = 3
        COL_APPOINTMENT = 4
        COL_END_TIME = 5
        COL_DROPOFF_TIME = 6
        COL_NOTES = 7
        COL_VOLUNTEER = 8
        COL_PICKUP = 0
        COL_ADDRESS = 0
        COL_PHONE = 0

    COL_TITLES = [
        (COL_DATE, 'Date'),
        (COL_PICKUP, 'Pick up'),
        (COL_APPOINTMENT, 'Appt. Time'),
        (COL_NAME, 'Name'),
        (COL_ADDRESS, 'Address'),
        (COL_PHONE, 'Phone #'),
        (COL_DESTINATION, 'Destination'),
        (COL_DRIVER, 'Driver'),
        (COL_VEHICLE, 'Vehicle'),
        (COL_START_MILES, 'Start Miles'),
        (COL_START_TIME, 'Start Time'),
        (COL_END_MILES, 'End Miles'),
        (COL_END_TIME, 'End Time'),
        (COL_NOTES, 'Notes'),
        (COL_TRIPTYPE, 'Trip Type / Tags'),
        (COL_PASSENGER, 'Passenger on vehicle?'),
        (COL_FARE, 'Fare'),
        (COL_MONEY, 'Money Collected'),
        (COL_ELDERLY, 'Elderly?'),
        (COL_AMBULATORY, 'Ambulatory?'),
        (COL_VOLUNTEER, 'Volunteer Driver'),
        (COL_DROPOFF_TIME, 'Time Difference'),
    ]

    COL_LAYOUT = []

    for i in range(0, len(COL_TITLES)):
        if COL_TITLES[i][0] != 0:
            COL_LAYOUT.append((COL_TITLES[i][0], i))

    COL_LAYOUT = sorted(COL_LAYOUT, key=lambda x: x[0])

    for col in COL_LAYOUT:
        ws_results.cell(row_header, col[0], COL_TITLES[col[1]][1])

    for i in range(0, row_header + trip_count):
        row = i + 1

        # apply styles
        if row > row_header:
            trip = trips[i-1]
            fill_color = trip.get_driver_color()[0:6]

        for col in range(1, len(COL_LAYOUT)+1):
            if col == COL_NAME or col == COL_PHONE or col == COL_NOTES or col == COL_TRIPTYPE:
                ws_results.column_dimensions[get_column_letter(col)].width = style_colwidth_large
            elif col == COL_ADDRESS or col == COL_DESTINATION or col == COL_VOLUNTEER:
                ws_results.column_dimensions[get_column_letter(col)].width = style_colwidth_xlarge
            else:
                ws_results.column_dimensions[get_column_letter(col)].width = style_colwidth_normal

            ws_results.cell(row, col).border = style_border_normal
            ws_results.cell(row, col).font = style_font_normal

            if row == row_header:
                ws_results.cell(row, col).font = style_font_header
                ws_results.cell(row, col).alignment = style_alignment_header
                ws_results.cell(row, col).fill = style_fill_header
            else:
                if COL_DATE != 0 and col == COL_DATE:
                    ws_results.cell(row, col).alignment = style_alignment_date

                ws_results.cell(row, col).fill = PatternFill(fill_type='solid', fgColor=fill_color)

                # text wrapping
                # TODO should all columns be wrapped?
                if COL_NOTES != 0:
                    ws_results.cell(row, COL_NOTES).alignment = style_alignment_wrapped
                if COL_NAME != 0:
                    ws_results.cell(row, COL_NAME).alignment = style_alignment_wrapped
                if COL_ADDRESS != 0:
                    ws_results.cell(row, COL_ADDRESS).alignment = style_alignment_wrapped
                if COL_DESTINATION != 0:
                    ws_results.cell(row, COL_DESTINATION).alignment = style_alignment_wrapped
                if COL_VOLUNTEER != 0:
                    ws_results.cell(row, COL_VOLUNTEER).alignment = style_alignment_wrapped
                if COL_TRIPTYPE != 0:
                    ws_results.cell(row, COL_TRIPTYPE).alignment = style_alignment_wrapped

                # number formats
                if COL_DATE != 0:
                    ws_results.cell(row, COL_DATE).number_format = 'mmm dd, yyyy'
                if COL_FARE != 0:
                    ws_results.cell(row, COL_FARE).number_format = '$0.00'
                if COL_MONEY != 0:
                    ws_results.cell(row, COL_MONEY).number_format = '$0.00'


        if row == row_header:
            continue

        if COL_DATE != 0:
            ws_results.cell(row, COL_DATE, trip.date)

        if COL_PICKUP != 0:
            ws_results.cell(row, COL_PICKUP, trip.pick_up_time)

        if COL_APPOINTMENT != 0:
            ws_results.cell(row, COL_APPOINTMENT, trip.appointment_time)

        if COL_NAME != 0:
            ws_results.cell(row, COL_NAME, trip.name)

        if COL_ADDRESS != 0:
            ws_results.cell(row, COL_ADDRESS, trip.address)

        if COL_PHONE != 0:
            phone_string = ''
            phone_string += trip.phone_home
            if trip.phone_home and trip.phone_cell:
                phone_string += ' / '
            phone_string += trip.phone_cell
            if trip.phone_alt and (trip.phone_home or trip.phone_cell):
                phone_string += ' / '
            phone_string += trip.phone_alt

            ws_results.cell(row, COL_PHONE, phone_string)

        if COL_DESTINATION != 0:
            ws_results.cell(row, COL_DESTINATION, trip.destination)

        if COL_DRIVER != 0 and trip.driver:
            ws_results.cell(row, COL_DRIVER, str(trip.driver))

        if COL_VEHICLE != 0 and trip.vehicle:
            ws_results.cell(row, COL_VEHICLE, str(trip.vehicle))

        if COL_START_MILES != 0:
            ws_results.cell(row, COL_START_MILES, trip.start_miles)

        if COL_START_TIME != 0:
            ws_results.cell(row, COL_START_TIME, trip.start_time)

        if COL_END_MILES != 0:
            ws_results.cell(row, COL_END_MILES, trip.end_miles)

        if COL_END_TIME != 0:
            ws_results.cell(row, COL_END_TIME, trip.end_time)

        if COL_NOTES != 0:
            ws_results.cell(row, COL_NOTES, trip.note)

        if COL_TRIPTYPE != 0:
            tags_string = ''
            if trip.trip_type:
                tags_string += str(trip.trip_type)
                if trip.tags:
                    tags_string += ' / '
            if trip.tags:
                tags_string += trip.tags
            ws_results.cell(row, COL_TRIPTYPE, tags_string)

        if COL_PASSENGER != 0:
            ws_results.cell(row, COL_PASSENGER, trip.passenger)

        if COL_FARE != 0:
            ws_results.cell(row, COL_FARE, trip.fare / 100)

        if COL_MONEY != 0:
            ws_results.cell(row, COL_MONEY, (trip.collected_cash + trip.collected_check) / 100)

        if COL_ELDERLY != 0:
            ws_results.cell(row, COL_ELDERLY, trip.elderly)

        if COL_AMBULATORY != 0:
            ws_results.cell(row, COL_AMBULATORY, trip.ambulatory)

        if COL_VOLUNTEER != 0 and trip.volunteer:
            ws_results.cell(row, COL_VOLUNTEER, trip.volunteer.verbose_name())

        if COL_DROPOFF_TIME != 0:
            ws_results.cell(row, COL_DROPOFF_TIME, trip.get_appt_dropoff_diff_xlsx())

    wb.save(filename=temp_file.name)

    return FileResponse(open(temp_file.name, 'rb'), filename='Transit_Search.xlsx', as_attachment=True)
