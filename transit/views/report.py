# Copyright © 2019-2023 Justin Jacobs
#
# This file is part of the Transit Log System.
#
# The Transit Log System is free software: you can redistribute it and/or modify it under the terms
# of the GNU General Public License as published by the Free Software Foundation,
# either version 3 of the License, or (at your option) any later version.
#
# The Transit Log System is distributed in the hope that it will be useful, but WITHOUT ANY
# WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS FOR A
# PARTICULAR PURPOSE.  See the GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License along with
# The Transit Log System.  If not, see http://www.gnu.org/licenses/

import datetime
import tempfile

from django.http import HttpResponseRedirect
from django.http import FileResponse
from django.shortcuts import render
from django.urls import reverse
from time import perf_counter

from transit.models import Driver, Vehicle, Trip, Shift, TripType, Client, ClientPayment, Tag, Destination
from transit.forms import DatePickerForm, DateRangePickerForm

from django.contrib.auth.decorators import permission_required

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

from transit.common.util import *

class Report():
    service_mile_warning_threshold = 1000

    RIDER_ELDERLY_AMBULATORY = 0
    RIDER_ELDERLY_NONAMBULATORY = 1
    RIDER_NONELDERLY_AMBULATORY = 2
    RIDER_NONELDERLY_NONAMBULATORY = 3
    RIDER_UNKNOWN = 4
    RIDER_TOTAL = 5
    RIDER_STAFF = 6
    RIDER_TOTAL_WITH_STAFF = 7

    class Money():
        def __init__(self, default_value=0):
            self.value = default_value
        def __add__(self, other):
            return Report.Money(self.value + other.value)
        def __sub__(self, other):
            return Report.Money(self.value - other.value)
        def __str__(self):
            return int_to_money_string(self.value, show_currency=True)
        def to_float(self):
            return float(self.value) / 100

    class ValueString():
        def __init__(self):
            self.value = 0
            self.string = ''
            self.is_string_valid = True
        def __lt__(self, other):
            return self.value < other.value
        def __gt__(self, other):
            return self.value > other.value
        def __le__(self, other):
            return self.value <= other.value
        def __ge__(self, other):
            return self.value >= other.value
        def __eq__(self, other):
            return self.value == other.value
        def __ne__(self, other):
            return self.value != other.value
        def empty(self):
            return self.string == ''

    class Mileage(ValueString):
        def __add__(self, other):
            r = Report.Mileage()
            r.value = self.value + other.value
            r.is_string_valid = False
            return r
        def __sub__(self, other):
            r = Report.Mileage()
            r.value = self.value - other.value
            r.is_string_valid = False
            return r
        def __str__(self):
            if not self.is_string_valid:
                self.string = f"{self.value:.1f}"
                self.is_string_valid = True
            return self.string
        def setFromString(self, string):
            try:
                self.value = float(string)
            except:
                self.value = 0
                self.string = ''
                self.is_string_valid = True
                return 1
            self.string = string
            self.is_string_valid = True
            return 0
        def mergeStrings(self, base, suffix):
            merged = ''
            if len(suffix) < len(base):
                merged = base[0:len(base) - len(suffix)] + suffix
            else:
                merged = suffix
            if self.setFromString(merged) != 0:
                return 1
            else:
                return 0

    class Time(ValueString):
        # TODO is this an acceptable "fallback" value? Does it matter? This is a worst case anyway...
        fallback_value = datetime.datetime(year=1900, month=1, day=1, hour=8, minute=0)
        fallback_string = fallback_value.strftime('%I:%M %p')

        def __str__(self):
            if not self.is_string_valid:
                self.is_string_valid = True
                if self.value != 0:
                    self.string = self.value.strftime('%I:%M %p')
            return self.string
        def setFromString(self, string, use_fallback=True):
            try:
                self.value = datetime.datetime.strptime(string, '%I:%M %p')
            except:
                if use_fallback:
                    self.value = Report.Time.fallback_value
                    self.string = Report.Time.fallback_string
                else:
                    self.value = 0
                    self.string = ''
                self.is_string_valid = True
                return 1
            self.string = string
            self.is_string_valid = True
            return 0

    class Fuel(ValueString):
        def __str__(self):
            if not self.is_string_valid:
                self.string = f"{self.value:.1f}"
                self.is_string_valid = True
            return self.string
        def setFromString(self, string):
            if string == '':
                value = 0
            else:
                try:
                    self.value = float(string)
                except:
                    self.value = 0
                    self.string = ''
                    self.is_string_valid = True
                    return 1
            self.string = string
            self.is_string_valid = True
            return 0

    class TripCount():
        def __init__(self):
            self.passenger = 0
            self.no_passenger = 0
            self.total = 0
        def __add__(self, other):
            r = Report.TripCount()
            r.passenger = self.passenger + other.passenger
            r.no_passenger = self.no_passenger + other.no_passenger
            r.total = self.total + other.total
            return r
        def __sub__(self, other):
            r = Report.TripCount()
            r.passenger = self.passenger - other.passenger
            r.no_passenger = self.no_passenger - other.no_passenger
            r.total = self.total - other.total
            return r
        def addTrips(self, value, is_passenger):
            if is_passenger == False:
                self.no_passenger += value
            else:
                self.passenger += value
            self.total = self.passenger + self.no_passenger
        def setTrips(self, value, is_passenger):
            if is_passenger == False:
                self.no_passenger = value
            else:
                self.passenger = value
            self.total = self.passenger + self.no_passenger
        def addTripsFromTripCount(self, value):
            self.passenger += value.passenger
            self.no_passenger += value.no_passenger
            self.total = self.passenger + self.no_passenger

    class ReportShift():
        def __init__(self):
            self.shift = None
            self.start_miles = Report.Mileage()
            self.start_time = Report.Time()
            self.end_miles = Report.Mileage()
            self.end_time = Report.Time()
            self.fuel = Report.Fuel()
            self.start_trip = None
            self.end_trip = None

    class ReportTrip():
        def __init__(self):
            self.trip = None
            self.shift = None
            self.start_miles = Report.Mileage()
            self.start_time = Report.Time()
            self.end_miles = Report.Mileage()
            self.end_time = Report.Time()
            self.trip_type = None
            self.tags = []
            self.collected_cash = Report.Money(0)
            self.collected_check = Report.Money(0)
            self.other_employment = False

    class ReportDay():
        query_vehicles = ()
        query_drivers = ()

        def __init__(self):
            self.date = None
            self.shifts = []
            self.trips = []
            self.all = Report.ReportSummary()
            self.collected_cash = Report.Money(0)
            self.collected_check = Report.Money(0)
            self.paid_cash = Report.Money(0)
            self.paid_check = Report.Money(0)
            self.total_payments = Report.Money(0)
            self.total_fares = Report.Money(0)
            self.by_vehicle = [None] * len(self.query_vehicles)
            self.by_driver = [None] * len(self.query_drivers)

        def hasVehicleInShift(self, vehicle = None):
            for i in self.shifts:
                if (vehicle and i.shift and i.shift.vehicle == vehicle) or (vehicle == None and i.shift and i.shift.vehicle != None):
                    return True
            return False

        def hasDriverInShift(self, driver = None):
            for i in self.shifts:
                if (driver and i.shift and i.shift.driver == driver) or (driver == None and i.shift and i.shift.driver != None):
                    return True
            return False

    class ReportSummary():
        query_triptypes = ()
        query_tags = ()

        # summary types
        TYPE_NORMAL = 0
        TYPE_LOGGED = 1
        TYPE_NONLOGGED = 2

        def __init__(self):
            self.service_miles = 0
            self.service_hours = 0
            self.deadhead_miles = 0
            self.deadhead_hours = 0
            self.total_miles = 0
            self.total_hours = 0
            self.pmt = 0
            self.fuel = 0
            self.trip_types = {}
            self.trip_types_unknown = Report.TripCount()
            self.trip_types_total = Report.TripCount()
            self.tags = {}
            self.collected_cash = Report.Money(0)
            self.collected_check = Report.Money(0)
            self.total_collected_money = Report.Money(0)
            self.other_employment = Report.TripCount()

            # flag to mark if this summary can be added to the grand total (i.e. "All Vehicles")
            self.type = Report.ReportSummary.TYPE_NORMAL

            for i in self.query_triptypes:
                self.trip_types[i] = Report.TripCount()

        def __add__(self, other):
            r = self
            if not (r.type == Report.ReportSummary.TYPE_LOGGED and other.type == Report.ReportSummary.TYPE_NONLOGGED):
                r.service_miles += other.service_miles
                r.service_hours += other.service_hours
                r.deadhead_miles += other.deadhead_miles
                r.deadhead_hours += other.deadhead_hours
                r.total_miles += other.total_miles
                r.total_hours += other.total_hours
                r.pmt += other.pmt
                r.fuel += other.fuel
            for i in self.query_triptypes:
                r.trip_types[i] += other.trip_types[i]
            r.trip_types_unknown += other.trip_types_unknown
            r.trip_types_total += other.trip_types_total
            for i in other.tags:
                if i in r.tags:
                    r.tags[i] += other.tags[i]
                else:
                    r.tags[i] = other.tags[i]
            r.collected_cash += other.collected_cash
            r.collected_check += other.collected_check
            r.total_collected_money += other.total_collected_money
            r.other_employment += other.other_employment
            return r

    class ReportOutputVehicles():
        def __init__(self):
            self.vehicle = None
            self.start_miles = Report.Mileage()
            self.end_miles = Report.Mileage()
            self.total_miles = Report.Mileage()
            self.days = []
            self.totals = Report.ReportSummary()

    class ReportOutputDrivers():
        def __init__(self):
            self.driver = None
            self.days = []
            self.totals = Report.ReportSummary()

    class ReportErrors():
        SHIFT_INCOMPLETE = 0
        TRIP_INCOMPLETE = 1
        TRIP_NO_SHIFT = 2
        TRIP_MILES_OOB = 3
        TRIP_TIME_OOB = 4
        SHIFT_PARSE = 5
        TRIP_PARSE = 6
        SHIFT_MILES_LESS = 7
        SHIFT_TIME_LESS = 8
        TRIP_MILES_LESS = 9
        TRIP_TIME_LESS = 10
        SHIFT_SERVICE_MILE_THRESHOLD = 11
        TRIP_EMPTY = 12

        def __init__(self):
            self.errors = []

        def add(self, date, daily_log_shift, error_code, error_shift=None, error_trip=None):
            if daily_log_shift == None or (error_shift and daily_log_shift == error_shift.id):
                self.errors.append({'date':date, 'error_code':error_code, 'error_shift': error_shift, 'error_trip': error_trip, 'error_msg': self.getErrorMsg(error_code, error_shift, error_trip)})

        def getErrorMsg(self, error_code, error_shift=None, error_trip=None):
            if (error_shift):
                shift_start_miles = error_shift.start_miles
                shift_end_miles = error_shift.end_miles
                shift_start_time = error_shift.start_time
                shift_end_time = error_shift.end_time
            else:
                shift_start_miles = shift_end_miles = shift_start_time = shift_end_time = ''

            if (error_trip):
                if (error_shift):
                    trip_start_miles = shift_start_miles[:len(shift_start_miles)-len(error_trip.start_miles)] + error_trip.start_miles
                    trip_end_miles = shift_start_miles[:len(shift_start_miles)-len(error_trip.end_miles)] + error_trip.end_miles
                else:
                    trip_start_miles = error_trip.start_miles
                    trip_end_miles = error_trip.end_miles
                trip_start_time = error_trip.start_time
                trip_end_time = error_trip.end_time
            else:
                trip_start_miles = trip_end_miles = trip_start_time = trip_end_time = ''

            if error_code == self.SHIFT_INCOMPLETE:
                return 'Shift contains partial log data.'

            elif error_code == self.TRIP_INCOMPLETE:
                return 'Trip contains partial log data.'

            elif error_code == self.TRIP_NO_SHIFT:
                return 'Trip does not have a matching Shift.'

            elif error_code == self.TRIP_MILES_OOB:
                return 'Trip mileage ( ' + trip_start_miles + ' - ' + trip_end_miles + ' ) is not within Shift mileage ( ' + shift_start_miles + ' - ' + shift_end_miles + ' ).'

            elif error_code == self.TRIP_TIME_OOB:
                return 'Trip time ( ' + trip_start_time + ' - ' + trip_end_time + ' ) is not within Shift time ( ' + shift_start_time + ' - ' + shift_end_time + ' ).'

            elif error_code == self.SHIFT_PARSE:
                return 'Unable to parse Shift data.'

            elif error_code == self.TRIP_PARSE:
                return 'Unable to parse Trip data.'

            elif error_code == self.SHIFT_MILES_LESS:
                return 'Shift start mileage ( ' + shift_start_miles + ' ) is greater than Shift end mileage ( ' + shift_end_miles + ' ).'

            elif error_code == self.SHIFT_TIME_LESS:
                return 'Shift start time ( ' + shift_start_time + ' ) is later than Shift end time ( ' + shift_end_time + ' ).'

            elif error_code == self.TRIP_MILES_LESS:
                return 'Trip start mileage ( ' + trip_start_miles + ' ) is greater than Trip end mileage ( ' + trip_end_miles + ' ).'

            elif error_code == self.TRIP_TIME_LESS:
                return 'Trip start time ( ' + trip_start_time + ' ) is later than Trip end time ( ' + trip_end_time + ' ).'

            elif error_code == self.SHIFT_SERVICE_MILE_THRESHOLD:
                return 'Shift mileage seems abnormally large.'

            elif error_code == self.TRIP_EMPTY:
                return 'Trip has no log data, but matches a completed Shift.'

            else:
                return 'Unknown error'

    class FrequentDestination():
        def __init__(self):
            self.address = None
            self.trips = Report.TripCount()
            self.avg_mileage = 0
        def __lt__(self, other):
            return self.trips.total < other.trips.total
        def averageMiles(self, miles):
            if self.avg_mileage == 0:
                self.avg_mileage = miles
            else:
                self.avg_mileage = (self.avg_mileage + miles) / 2

    class UniqueRiderSummary():
        class Rider():
            def __init__(self, name):
                self.name = name
                self.client_id = None
                self.trips = Report.TripCount()
                self.elderly = None
                self.ambulatory = None
                self.collected_cash = Report.Money(0)
                self.collected_check = Report.Money(0)
                self.paid_cash = Report.Money(0)
                self.paid_check = Report.Money(0)
                self.total_payments = Report.Money(0)
                self.total_fares = Report.Money(0)
                self.total_owed = Report.Money(0)
                self.staff = False
                self.trips_no_show = Report.TripCount()
                self.trips_canceled_late = Report.TripCount()
                self.trips_canceled_very_late = Report.TripCount()
            def __lt__(self, other):
                return self.name < other.name

        def __init__(self):
            self.names = {}
            self.by_individuals = [Report.TripCount() for i in range(8)]
            self.by_trips = [Report.TripCount() for i in range(8)]
            # fares & payments totals
            self.total_collected_cash = Report.Money(0)
            self.total_collected_check = Report.Money(0)
            self.total_paid_cash = Report.Money(0)
            self.total_paid_check = Report.Money(0)
            self.total_total_payments = Report.Money(0)
            self.total_total_fares = Report.Money(0)
            self.total_total_owed = Report.Money(0)

    class ReportPayment():
        def __init__(self):
            self.date = None
            self.id = None
            self.client = None
            self.cash = Report.Money(0)
            self.check = Report.Money(0)

    def __init__(self):
        Report.ReportSummary.query_triptypes = TripType.objects.filter(is_trip_counted=True)
        Report.ReportSummary.query_tags = Tag.objects.all()

        self.report_all = []
        self.vehicle_reports = []
        self.driver_reports = []
        self.driver_reports_total = Report.ReportOutputDrivers()
        self.all_vehicles = Report.ReportSummary()
        self.unique_riders = Report.UniqueRiderSummary()
        self.money_trips = []
        self.money_trips_summary = Report.ReportSummary()
        self.money_payments = []
        self.money_payments_summary = Report.ReportPayment()
        self.frequent_destinations = {}
        self.report_errors = Report.ReportErrors()
        self.filtered_vehicles = None
        self.filtered_drivers = None
        self.total_vehicle_days_of_service = 0
        self.total_vehicle_mileage = Report.Mileage()
        self.total_money = Report.Money(0)
        self.total_odometer_miles = Report.Mileage()

        self.perf_database = 0
        self.perf_processing = 0

        self.weekday_totals = []
        for i in range(0, 7):
            self.weekday_totals.append(Report.TripCount())

    def load(self, date_start, date_end, daily_log_shift=None, driver_id=None, client_name=None, filter_by_money=False):
        if date_start != date_end:
            daily_log_shift = None

        # for filter() bounds
        date_end_plus_one = date_end + datetime.timedelta(days=1)

        perf_start = perf_counter()
        # refresh related fields
        # without this, select_related can fail if new entries are used
        all_drivers = Driver.objects.all()
        all_vehicles = Vehicle.objects.all()
        all_triptypes = TripType.objects.all()

        self.filtered_vehicles = all_vehicles

        if driver_id == None:
            self.filtered_drivers = all_drivers.filter(is_logged=True)
        else:
            self.filtered_drivers = all_drivers.filter(id=driver_id)

        # also cache destinations
        all_destinations = Destination.objects.all()
        destination_dict = {}
        for i in all_destinations:
            destination_dict[i.address] = i

        all_clients = Client.objects.all()
        if client_name != None:
            all_clients = all_clients.filter(name=client_name)

        client_dict = {}
        for i in all_clients:
            client_dict[i.name] = i

        # store our database lookups
        all_shifts = Shift.objects.filter(date__gte=date_start, date__lt=date_end_plus_one, status=Shift.STATUS_NORMAL)
        all_shifts = all_shifts.select_related('driver').select_related('vehicle')

        all_trips = Trip.objects.filter(date__gte=date_start, date__lt=date_end_plus_one, format=Trip.FORMAT_NORMAL)
        if filter_by_money:
            all_trips = all_trips.exclude(fare=0, collected_cash=0, collected_check=0)
        if client_name != None:
            all_trips = all_trips.filter(name=client_name)
        all_trips = all_trips.select_related('driver').select_related('vehicle').select_related('trip_type')

        all_client_payments = ClientPayment.objects.filter(date_paid__gte=date_start, date_paid__lt=date_end_plus_one)
        all_client_payments = all_client_payments.select_related('parent')
        if client_name != None:
            all_client_payments = all_client_payments.filter(parent__name=client_name)
        self.perf_database = perf_counter() - perf_start

        perf_start = perf_counter()
        Report.ReportDay.query_drivers = list(self.filtered_drivers)
        Report.ReportDay.query_vehicles = list(all_vehicles)

        all_dates = []

        all_dates_dict = {}
        for i in all_shifts:
            if i.date not in all_dates_dict:
                all_dates_dict[i.date] = [[], [], []]
            all_dates_dict[i.date][0].append(i)

        for i in all_trips:
            if i.date not in all_dates_dict:
                all_dates_dict[i.date] = [[], [], []]
            all_dates_dict[i.date][1].append(i)

        for i in all_client_payments:
            if i.date_paid not in all_dates_dict:
                all_dates_dict[i.date_paid] = [[], [], []]
            all_dates_dict[i.date_paid][2].append(i)

        all_dates_dict = {k: v for k, v in sorted(all_dates_dict.items(), key=lambda x: x[0])}

        self.all_vehicles = Report.ReportSummary()

        # make sure our pre-defined tags get placed first in the final list
        for i in Report.ReportSummary.query_tags:
            self.all_vehicles.tags[i.name] = Report.TripCount()

        self.vehicle_reports = []
        for vehicle in self.filtered_vehicles:
            vehicle_report = Report.ReportOutputVehicles()
            vehicle_report.vehicle = vehicle
            self.vehicle_reports.append(vehicle_report)

        self.driver_reports = []
        for driver in self.filtered_drivers:
            driver_report = Report.ReportOutputDrivers()
            driver_report.driver = driver
            self.driver_reports.append(driver_report)

        def UniqueRiderInit(trip):
            if trip.name not in self.unique_riders.names:
                self.unique_riders.names[trip.name] = Report.UniqueRiderSummary.Rider(trip.name)

            rider = self.unique_riders.names[trip.name]
            rider.elderly = trip.elderly
            rider.ambulatory = trip.ambulatory

            if trip.name in client_dict:
                client = client_dict[trip.name]

                rider.client_id = client.id
                rider.staff = client.staff

                if trip.elderly == None or trip.ambulatory == None:
                    # try to get info from Clients
                    if rider.elderly == None:
                        rider.elderly = client.elderly
                    if rider.ambulatory == None:
                        rider.ambulatory = client.ambulatory

            return rider

        for day_date in all_dates_dict:
            report_day = Report.ReportDay()
            report_day.all.type = Report.ReportSummary.TYPE_LOGGED
            report_day.date = day_date

            for i in all_dates_dict[day_date][0]:
                if driver_id == None:
                    if i.driver and not i.driver.is_logged:
                        # skip non-logged drivers
                        continue
                elif i.driver:
                    if i.driver.id != driver_id:
                        continue

                log_status = i.check_log()
                if log_status != Shift.LOG_COMPLETE or i.driver == None or i.vehicle == None:
                    # skip incomplete shift
                    if log_status == Shift.LOG_INCOMPLETE:
                        self.report_errors.add(day_date, daily_log_shift, self.report_errors.SHIFT_INCOMPLETE, error_shift=i)
                    continue

                report_shift = Report.ReportShift()
                report_shift.shift = i

                if report_shift.start_miles.setFromString(i.start_miles) != 0:
                    self.report_errors.add(day_date, daily_log_shift, self.report_errors.SHIFT_PARSE, error_shift=i)

                if report_shift.start_time.setFromString(i.start_time) != 0:
                    self.report_errors.add(day_date, daily_log_shift, self.report_errors.SHIFT_PARSE, error_shift=i)

                if report_shift.end_miles.setFromString(i.end_miles) != 0:
                    self.report_errors.add(day_date, daily_log_shift, self.report_errors.SHIFT_PARSE, error_shift=i)

                if report_shift.end_time.setFromString(i.end_time) != 0:
                    self.report_errors.add(day_date, daily_log_shift, self.report_errors.SHIFT_PARSE, error_shift=i)

                if report_shift.fuel.setFromString(i.fuel) != 0:
                    self.report_errors.add(day_date, daily_log_shift, self.report_errors.SHIFT_PARSE, error_shift=i)

                if report_shift.start_miles > report_shift.end_miles:
                    self.report_errors.add(day_date, daily_log_shift, self.report_errors.SHIFT_MILES_LESS, error_shift=i)
                    report_shift.end_miles = report_shift.start_miles

                if report_shift.start_time > report_shift.end_time:
                    self.report_errors.add(day_date, daily_log_shift, self.report_errors.SHIFT_TIME_LESS, error_shift=i)
                    report_shift.end_time = report_shift.start_time

                if report_shift.end_miles.value - report_shift.start_miles.value > Report.service_mile_warning_threshold:
                    self.report_errors.add(day_date, daily_log_shift, self.report_errors.SHIFT_SERVICE_MILE_THRESHOLD, error_shift=i)

                report_day.shifts.append(report_shift)

            for i in all_dates_dict[day_date][1]:
                if i.status == Trip.STATUS_NO_SHOW:
                    rider = UniqueRiderInit(i)
                    rider.trips_no_show.addTrips(1, i.passenger)
                    continue
                elif i.status == Trip.STATUS_CANCELED:
                    rider = UniqueRiderInit(i)
                    cancel_status = i.check_cancel_date()
                    if cancel_status == 3:
                        rider.trips_canceled_very_late.addTrips(1, i.passenger)
                    elif cancel_status == 2:
                        rider.trips_canceled_late.addTrips(1, i.passenger)
                    continue

                if not i.driver and not i.vehicle:
                    continue

                log_status = i.check_log()

                if driver_id == None:
                    if i.driver and not i.driver.is_logged:
                        # skip non-logged drivers
                        continue
                elif i.driver:
                    if i.driver.id != driver_id:
                        continue

                # skip trip with no driver/vehicle
                if (i.driver and not i.vehicle) or (not i.driver and i.vehicle):
                    self.report_errors.add(day_date, daily_log_shift, self.report_errors.TRIP_INCOMPLETE, error_trip=i)
                    continue

                report_trip = Report.ReportTrip()
                report_trip.trip = i

                matched_shifts = []

                # find shift attempt 1: match driver and vehicle
                for j in range(0, len(report_day.shifts)):
                    if report_day.shifts[j].shift and i.driver == report_day.shifts[j].shift.driver and i.vehicle == report_day.shifts[j].shift.vehicle:
                        matched_shifts.append(j)

                # find shift attempt 2: match vehicle
                if len(matched_shifts) == 0:
                    for j in range(0, len(report_day.shifts)):
                        if report_day.shifts[j].shift and i.vehicle == report_day.shifts[j].shift.vehicle:
                            matched_shifts.append(j)

                # if there are multiple matching shifts, find the first one that the trip mileage fits in to
                # TODO should time also be considered?
                if len(matched_shifts) > 1 and log_status == Trip.LOG_COMPLETE:
                    for match in matched_shifts:
                        test_shift = report_day.shifts[match]

                        start_miles = Report.Mileage()
                        start_miles.mergeStrings(str(test_shift.start_miles), i.start_miles)
                        if start_miles < test_shift.start_miles or start_miles > test_shift.end_miles:
                            start_miles.mergeStrings(str(test_shift.end_miles), i.start_miles)
                            if start_miles < test_shift.start_miles or start_miles > test_shift.end_miles:
                                continue

                        end_miles = Report.Mileage()
                        end_miles.mergeStrings(str(test_shift.start_miles), i.end_miles)
                        if end_miles < test_shift.start_miles or end_miles > test_shift.end_miles:
                            end_miles.mergeStrings(str(test_shift.end_miles), i.end_miles)
                            if end_miles < test_shift.start_miles or end_miles > test_shift.end_miles:
                                continue

                        report_trip.shift = match
                        break
                elif len(matched_shifts) == 1:
                    report_trip.shift = matched_shifts[0]

                # find shift attempt 3: create a dummy shift (or skip the trip if the vehicle is logged)
                if report_trip.shift == None:
                    if i.vehicle.is_logged:
                        if log_status != Trip.LOG_EMPTY:
                            self.report_errors.add(day_date, daily_log_shift, self.report_errors.TRIP_NO_SHIFT, error_trip=i)
                        continue
                    else:
                        dummy_shift = Report.ReportShift()
                        dummy_shift.shift = Shift()
                        dummy_shift.driver = dummy_shift.shift.driver = i.driver
                        dummy_shift.vehicle = dummy_shift.shift.vehicle = i.vehicle
                        report_day.shifts.append(dummy_shift)
                        report_trip.shift = len(report_day.shifts)-1

                # TODO could report_trip.shift be None here?
                shift = report_day.shifts[report_trip.shift]

                # skip incomplete trip (logged vehicles only)
                if i.vehicle.is_logged and log_status == Trip.LOG_INCOMPLETE:
                    self.report_errors.add(day_date, daily_log_shift, self.report_errors.TRIP_INCOMPLETE, error_shift=shift.shift, error_trip=i)
                    continue
                elif i.vehicle.is_logged and log_status == Trip.LOG_EMPTY and shift.shift.check_log() == Shift.LOG_COMPLETE:
                    self.report_errors.add(day_date, daily_log_shift, self.report_errors.TRIP_EMPTY, error_shift=shift.shift, error_trip=i)
                    continue

                if log_status == Trip.LOG_COMPLETE and report_day.shifts[report_trip.shift].shift.check_log() == Shift.LOG_COMPLETE:
                    parse_error = False

                    if report_trip.start_miles.mergeStrings(str(shift.start_miles), i.start_miles) != 0:
                        parse_error = True

                    if report_trip.start_time.setFromString(i.start_time) != 0:
                        parse_error = True

                    if report_trip.end_miles.mergeStrings(str(shift.start_miles), i.end_miles) != 0:
                        parse_error = True

                    if report_trip.end_time.setFromString(i.end_time) != 0:
                        parse_error = True

                    if parse_error:
                        self.report_errors.add(day_date, daily_log_shift, self.report_errors.TRIP_PARSE, error_shift=shift.shift, error_trip=i)

                    # check for trip errors
                    # TODO the above execptions add errors to the report without adding the trip to the various trip counts
                    # BUT, the below errors still count the trips where possible. Not sure which is preferable, but it's probably bad that both behaviors exist?
                    # NOTE For mileage, we also check for validity by comparing against merges of shift *end* mileage and trip mileage.

                    # mileage error checking
                    if report_trip.start_miles < shift.start_miles or report_trip.start_miles > shift.end_miles or report_trip.end_miles < shift.start_miles or report_trip.end_miles > shift.end_miles:
                        # out-of-bounds
                        mile_error = False

                        # start mileage
                        if report_trip.start_miles < shift.start_miles or report_trip.start_miles > shift.end_miles:
                            mile_error = False
                            if report_trip.start_miles.mergeStrings(str(shift.end_miles), i.start_miles) == 0:
                                if report_trip.start_miles < shift.start_miles or report_trip.start_miles > shift.end_miles:
                                    mile_error = True
                            else:
                                mile_error = True
                            if mile_error:
                                report_trip.start_miles = shift.start_miles

                        # end mileage
                        if report_trip.end_miles < shift.start_miles or report_trip.end_miles > shift.end_miles:
                            mile_error = False
                            if report_trip.end_miles.mergeStrings(str(shift.end_miles), i.end_miles) == 0:
                                if report_trip.end_miles < shift.start_miles or report_trip.end_miles > shift.end_miles:
                                    mile_error = True
                            else:
                                mile_error = True
                            if mile_error:
                                report_trip.end_miles = shift.end_miles

                        if mile_error:
                            self.report_errors.add(day_date, daily_log_shift, self.report_errors.TRIP_MILES_OOB, error_shift=shift.shift, error_trip=i)
                    # end > start
                    elif report_trip.start_miles > report_trip.end_miles:
                        mile_error = False
                        if report_trip.end_miles.mergeStrings(str(shift.end_miles), i.end_miles) == 0:
                            if report_trip.start_miles > report_trip.end_miles or report_trip.end_miles > shift.end_miles:
                                mile_error = True
                        else:
                            mile_error = True
                        if mile_error:
                            report_trip.start_miles = shift.start_miles
                            report_trip.end_miles = shift.start_miles
                            self.report_errors.add(day_date, daily_log_shift, self.report_errors.TRIP_MILES_LESS, error_shift=shift.shift, error_trip=i)

                    # time error checking
                    if report_trip.start_time and shift.start_time and report_trip.end_time and shift.end_time:
                        # out-of-bounds
                        if report_trip.start_time < shift.start_time or report_trip.end_time > shift.end_time:
                            self.report_errors.add(day_date, daily_log_shift, self.report_errors.TRIP_TIME_OOB, error_shift=shift.shift, error_trip=i)
                            if report_trip.start_time < shift.start_time:
                                report_trip.start_time = shift.start_time
                            if report_trip.end_time > shift.end_time:
                                report_trip.end_time = shift.end_time
                        # end > start
                        elif report_trip.start_time > report_trip.end_time:
                            self.report_errors.add(day_date, daily_log_shift, self.report_errors.TRIP_TIME_LESS, error_shift=shift.shift, error_trip=i)
                            report_trip.start_time = shift.start_time
                            report_trip.end_time = shift.start_time

                report_trip.trip_type = i.trip_type
                report_trip.collected_cash = Report.Money(i.collected_cash)
                report_trip.collected_check = Report.Money(i.collected_check)

                report_day.collected_cash += report_trip.collected_cash
                report_day.collected_check += report_trip.collected_check
                report_day.total_payments += report_trip.collected_cash + report_trip.collected_check

                report_day.total_fares += Report.Money(i.fare)

                report_trip.other_employment = i.check_tag('Employment')

                report_day.trips.append(report_trip)

                if i.trip_type and i.trip_type.is_trip_counted:
                    self.weekday_totals[day_date.weekday()].addTrips(1, i.passenger)

                if log_status == Trip.LOG_COMPLETE:
                    if shift.start_trip == None or report_trip.start_miles < report_day.trips[shift.start_trip].start_miles:
                        report_day.shifts[report_trip.shift].start_trip = len(report_day.trips) - 1;

                    if shift.end_trip == None or report_trip.end_miles > report_day.trips[shift.end_trip].end_miles:
                        report_day.shifts[report_trip.shift].end_trip = len(report_day.trips) - 1;


                if daily_log_shift == None or (daily_log_shift != None and daily_log_shift == shift.shift.id):
                    # add money trip
                    if i.collected_cash > 0 or i.collected_check > 0:
                        self.money_trips.append(report_trip)
                        self.money_trips_summary.collected_cash += report_trip.collected_cash
                        self.money_trips_summary.collected_check += report_trip.collected_check

                    # add unique rider
                    rider = UniqueRiderInit(i)

                    rider.trips.addTrips(1, i.passenger)
                    rider.total_fares += Report.Money(i.fare)
                    rider.collected_cash += Report.Money(i.collected_cash)
                    rider.collected_check += Report.Money(i.collected_check)

                if i.tags != "":
                    report_trip.tags = i.get_tag_list()

                # add destination to frequent destinations
                if i.destination in destination_dict:
                    if i.destination not in self.frequent_destinations:
                        self.frequent_destinations[i.destination] = Report.FrequentDestination()
                        self.frequent_destinations[i.destination].address = i.destination
                    self.frequent_destinations[i.destination].trips.addTrips(1, i.passenger)
                    if log_status == Trip.LOG_COMPLETE:
                        self.frequent_destinations[i.destination].averageMiles(report_trip.end_miles.value - report_trip.start_miles.value)

            # handle payments from Clients that didn't ride (so far)
            for i in all_dates_dict[day_date][2]:
                # if client_name and i.parent.name != client_name:
                #     continue

                report_day.paid_cash += Report.Money(i.money_cash)
                report_day.paid_check += Report.Money(i.money_check)

                if i.parent.name not in self.unique_riders.names:
                    self.unique_riders.names[i.parent.name] = Report.UniqueRiderSummary.Rider(i.parent.name)
                    rider = self.unique_riders.names[i.parent.name]
                    rider.elderly = i.parent.elderly
                    rider.ambulatory = i.parent.ambulatory
                    rider.client_id = i.parent.id
                else:
                    rider = self.unique_riders.names[i.parent.name]

                rider.paid_cash += Report.Money(i.money_cash)
                rider.paid_check += Report.Money(i.money_check)

                # create a summary of only non-driver payments
                payment = Report.ReportPayment()
                payment.date = i.date_paid
                payment.id = i.id
                payment.client = i.parent
                payment.cash = Report.Money(i.money_cash)
                payment.check = Report.Money(i.money_check)
                self.money_payments.append(payment)
                self.money_payments_summary.cash += payment.cash
                self.money_payments_summary.check += payment.check

            report_day.total_payments += report_day.paid_cash + report_day.paid_check

            for i in range(0, len(report_day.shifts)):
                shift = report_day.shifts[i]

                # create dummy trip when a shift has no trips
                if shift.start_trip == None and shift.end_trip == None:
                    rt = Report.ReportTrip()
                    rt.shift = i
                    rt.start_miles = shift.start_miles
                    rt.end_miles = shift.end_miles
                    rt.start_time = shift.start_time
                    rt.end_time = shift.end_time
                    report_day.trips.append(rt)
                    shift.start_trip = len(report_day.trips) - 1
                    shift.end_trip = shift.start_trip
                elif shift.start_trip == None or shift.end_trip == None:
                    # TODO create partial dummy trips?
                    continue

                if daily_log_shift != None and daily_log_shift != shift.shift.id:
                    continue
                
                service_miles = 0
                service_hours = 0
                deadhead_miles = 0
                deadhead_hours = 0
                if shift.shift.check_log() == Shift.LOG_COMPLETE:
                    service_miles = shift.end_miles.value - shift.start_miles.value
                    service_hours = (shift.end_time.value - shift.start_time.value).seconds / 60 / 60
                    deadhead_miles = (report_day.trips[shift.start_trip].start_miles.value - shift.start_miles.value) + (shift.end_miles.value - report_day.trips[shift.end_trip].end_miles.value)
                    deadhead_hours = ((report_day.trips[shift.start_trip].start_time.value - shift.start_time.value).seconds + (shift.end_time.value - report_day.trips[shift.end_trip].end_time.value).seconds) / 60 / 60

                # per-vehicle and per-driver logs
                vehicle_index = Report.getVehicleIndex(shift.shift.vehicle)
                driver_index = Report.getDriverIndex(shift.shift.driver)

                if not report_day.by_vehicle[vehicle_index]:
                    report_day.by_vehicle[vehicle_index] = Report.ReportSummary()

                if not report_day.by_driver[driver_index]:
                    report_day.by_driver[driver_index] = Report.ReportSummary()

                if not shift.shift.vehicle.is_logged:
                    report_day.by_vehicle[vehicle_index].type = Report.ReportSummary.TYPE_NONLOGGED

                report_day.by_vehicle[vehicle_index].service_miles += service_miles
                report_day.by_vehicle[vehicle_index].service_hours += service_hours
                report_day.by_vehicle[vehicle_index].deadhead_miles += deadhead_miles
                report_day.by_vehicle[vehicle_index].deadhead_hours += deadhead_hours
                report_day.by_vehicle[vehicle_index].total_miles += service_miles + deadhead_miles
                report_day.by_vehicle[vehicle_index].total_hours += service_hours + deadhead_hours
                report_day.by_vehicle[vehicle_index].fuel += shift.fuel.value

                report_day.by_driver[driver_index].service_miles += service_miles
                report_day.by_driver[driver_index].service_hours += service_hours
                report_day.by_driver[driver_index].deadhead_miles += deadhead_miles
                report_day.by_driver[driver_index].deadhead_hours += deadhead_hours
                report_day.by_driver[driver_index].total_miles += service_miles + deadhead_miles
                report_day.by_driver[driver_index].total_hours += service_hours + deadhead_hours
                report_day.by_driver[driver_index].fuel += shift.fuel.value
                for trip in report_day.trips:
                    if i != trip.shift:
                        continue
                    report_day.by_vehicle[vehicle_index].pmt += trip.end_miles.value - trip.start_miles.value
                    report_day.by_vehicle[vehicle_index].collected_cash += trip.collected_cash
                    report_day.by_vehicle[vehicle_index].collected_check += trip.collected_check
                    report_day.by_vehicle[vehicle_index].total_collected_money += (trip.collected_cash + trip.collected_check)
                    report_day.by_driver[driver_index].pmt += trip.end_miles.value - trip.start_miles.value
                    report_day.by_driver[driver_index].collected_cash += trip.collected_cash
                    report_day.by_driver[driver_index].collected_check += trip.collected_check
                    report_day.by_driver[driver_index].total_collected_money += (trip.collected_cash + trip.collected_check)
                    if trip.trip != None:
                        if trip.trip_type != None and trip.trip_type.is_trip_counted:
                            report_day.by_vehicle[vehicle_index].trip_types[trip.trip_type].addTrips(1, trip.trip.passenger)
                            report_day.by_vehicle[vehicle_index].trip_types_total.addTrips(1, trip.trip.passenger)
                            report_day.by_driver[driver_index].trip_types[trip.trip_type].addTrips(1, trip.trip.passenger)
                            report_day.by_driver[driver_index].trip_types_total.addTrips(1, trip.trip.passenger)
                        elif trip.trip_type == None:
                            report_day.by_vehicle[vehicle_index].trip_types_unknown.addTrips(1, trip.trip.passenger)
                            report_day.by_vehicle[vehicle_index].trip_types_total.addTrips(1, trip.trip.passenger)
                            report_day.by_driver[driver_index].trip_types_unknown.addTrips(1, trip.trip.passenger)
                            report_day.by_driver[driver_index].trip_types_total.addTrips(1, trip.trip.passenger)
                        if trip.other_employment:
                            report_day.by_vehicle[vehicle_index].other_employment.addTrips(1, trip.trip.passenger)
                            report_day.by_driver[driver_index].other_employment.addTrips(1, trip.trip.passenger)
                        for tag in trip.tags:
                            if tag in report_day.by_vehicle[vehicle_index].tags:
                                report_day.by_vehicle[vehicle_index].tags[tag].addTrips(1, trip.trip.passenger)
                            elif tag != '':
                                report_day.by_vehicle[vehicle_index].tags[tag] = Report.TripCount()
                                report_day.by_vehicle[vehicle_index].tags[tag].setTrips(1, trip.trip.passenger)
                            if tag in report_day.by_driver[driver_index].tags:
                                report_day.by_driver[driver_index].tags[tag].addTrips(1, trip.trip.passenger)
                            elif tag != '':
                                report_day.by_driver[driver_index].tags[tag] = Report.TripCount()
                                report_day.by_driver[driver_index].tags[tag].setTrips(1, trip.trip.passenger)

            for shift_vehicle in report_day.by_vehicle:
                if shift_vehicle:
                    report_day.all += shift_vehicle

            self.report_all.append(report_day)

            for vehicle_report in self.vehicle_reports:
                if report_day.hasVehicleInShift(vehicle_report.vehicle):
                    vehicle_index = Report.getVehicleIndex(vehicle_report.vehicle)
                    if report_day.by_vehicle[vehicle_index] != None:
                        vehicle_report.totals.type = report_day.by_vehicle[vehicle_index].type
                        vehicle_report.days.append({'date':report_day.date, 'data': report_day.by_vehicle[vehicle_index]})
                        vehicle_report.totals += report_day.by_vehicle[vehicle_index]
                        for shift_iter in report_day.shifts:
                            if shift_iter.shift and vehicle_report.vehicle.id == shift_iter.shift.vehicle.id:
                                if vehicle_report.start_miles.empty() or (not vehicle_report.start_miles.empty() and vehicle_report.start_miles > shift_iter.start_miles):
                                    vehicle_report.start_miles = shift_iter.start_miles
                                if vehicle_report.end_miles.empty() or (not vehicle_report.end_miles.empty() and vehicle_report.end_miles < shift_iter.end_miles):
                                    vehicle_report.end_miles = shift_iter.end_miles
            for driver_report in self.driver_reports:
                if report_day.hasDriverInShift(driver_report.driver):
                    driver_index = Report.getDriverIndex(driver_report.driver)
                    if report_day.by_driver[driver_index] != None:
                        driver_report.days.append({'date':report_day.date, 'data': report_day.by_driver[driver_index]})
                        driver_report.totals += report_day.by_driver[driver_index]

            if report_day.hasVehicleInShift():
                self.total_vehicle_days_of_service += 1


        cleaned_driver_reports = []
        for driver_report in self.driver_reports:
            if len(driver_report.days) > 0:
                cleaned_driver_reports.append(driver_report)
                self.driver_reports_total.totals += driver_report.totals
        self.driver_reports = cleaned_driver_reports

        self.all_vehicles.type = Report.ReportSummary.TYPE_LOGGED
        for vehicle_report in self.vehicle_reports:
            if vehicle_report.end_miles >= vehicle_report.start_miles:
                vehicle_report.total_miles = vehicle_report.end_miles - vehicle_report.start_miles
            self.total_vehicle_mileage += vehicle_report.total_miles
            self.all_vehicles += vehicle_report.totals

        for rider_name in self.unique_riders.names:
            rider = self.unique_riders.names[rider_name]
            # total elderly/ambulatory counts
            if rider.trips.total > 0:
                is_passenger = rider.trips.passenger > 0

                self.unique_riders.by_individuals[Report.RIDER_TOTAL_WITH_STAFF].addTrips(1, is_passenger)
                self.unique_riders.by_trips[Report.RIDER_TOTAL_WITH_STAFF].addTripsFromTripCount(rider.trips)

                if rider.staff:
                    self.unique_riders.by_individuals[Report.RIDER_STAFF].addTrips(1, is_passenger)
                    self.unique_riders.by_trips[Report.RIDER_STAFF].addTripsFromTripCount(rider.trips)
                else:
                    self.unique_riders.by_individuals[Report.RIDER_TOTAL].addTrips(1, is_passenger)
                    self.unique_riders.by_trips[Report.RIDER_TOTAL].addTripsFromTripCount(rider.trips)

                    if rider.elderly == None or rider.ambulatory == None:
                        self.unique_riders.by_individuals[Report.RIDER_UNKNOWN].addTrips(1, is_passenger)
                        self.unique_riders.by_trips[Report.RIDER_UNKNOWN].addTripsFromTripCount(rider.trips)
                    elif rider.elderly and rider.ambulatory:
                        self.unique_riders.by_individuals[Report.RIDER_ELDERLY_AMBULATORY].addTrips(1, is_passenger)
                        self.unique_riders.by_trips[Report.RIDER_ELDERLY_AMBULATORY].addTripsFromTripCount(rider.trips)
                    elif rider.elderly and not rider.ambulatory:
                        self.unique_riders.by_individuals[Report.RIDER_ELDERLY_NONAMBULATORY].addTrips(1, is_passenger)
                        self.unique_riders.by_trips[Report.RIDER_ELDERLY_NONAMBULATORY].addTripsFromTripCount(rider.trips)
                    elif not rider.elderly and rider.ambulatory:
                        self.unique_riders.by_individuals[Report.RIDER_NONELDERLY_AMBULATORY].addTrips(1, is_passenger)
                        self.unique_riders.by_trips[Report.RIDER_NONELDERLY_AMBULATORY].addTripsFromTripCount(rider.trips)
                    elif not rider.elderly and not rider.ambulatory:
                        self.unique_riders.by_individuals[Report.RIDER_NONELDERLY_NONAMBULATORY].addTrips(1, is_passenger)
                        self.unique_riders.by_trips[Report.RIDER_NONELDERLY_NONAMBULATORY].addTripsFromTripCount(rider.trips)

            # calculate total owed money
            rider.total_payments = rider.collected_cash + rider.collected_check + rider.paid_cash + rider.paid_check
            if rider.total_payments.value != 0 or rider.total_fares.value != 0:
                rider.total_owed = rider.total_fares - rider.total_payments
                if (rider.total_owed.value < 0):
                    rider.total_owed.value = 0

            self.unique_riders.total_collected_cash += rider.collected_cash
            self.unique_riders.total_collected_check += rider.collected_check
            self.unique_riders.total_paid_cash += rider.paid_cash
            self.unique_riders.total_paid_check += rider.paid_check
            self.unique_riders.total_total_payments += rider.total_payments
            self.unique_riders.total_total_fares += rider.total_fares
            self.unique_riders.total_total_owed += rider.total_owed

        # sort frequent destinations by total trips
        self.frequent_destinations = {k: v for k, v in sorted(self.frequent_destinations.items(), key=lambda x: x[1], reverse=True)}

        # sort unique riders by name
        self.unique_riders.names = {k: v for k, v in sorted(self.unique_riders.names.items(), key=lambda x: x[0])}

        self.total_money = self.all_vehicles.total_collected_money + self.money_payments_summary.cash + self.money_payments_summary.check

        for vehicle_report in self.vehicle_reports:
            if vehicle_report.totals.type == Report.ReportSummary.TYPE_NORMAL:
                self.total_odometer_miles += vehicle_report.total_miles

        self.perf_processing = perf_counter() - perf_start

    def getVehicleIndex(vehicle):
        return Report.ReportDay.query_vehicles.index(vehicle)

    def getDriverIndex(driver):
        return Report.ReportDay.query_drivers.index(driver)

@permission_required(['transit.view_trip', 'transit.view_shift'])
def report(request, start_year, start_month, start_day, end_year, end_month, end_day):
    return reportBase(request, None, start_year, start_month, start_day, end_year, end_month, end_day)

@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportDriver(request, driver_id, start_year, start_month, start_day, end_year, end_month, end_day):
    return reportBase(request, driver_id, start_year, start_month, start_day, end_year, end_month, end_day)

@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportBase(request, driver_id, start_year, start_month, start_day, end_year, end_month, end_day):
    date_start = datetime.date(start_year, start_month, start_day)
    date_end = datetime.date(end_year, end_month, end_day)

    show_daily_data = request.session.get('report_show_daily_data', False)
    request.session['report_show_daily_data'] = show_daily_data

    if date_start > date_end:
        swap_date = date_start
        date_start = date_end
        date_end = swap_date

    if request.method == 'POST':
        date_picker = DatePickerForm(request.POST)
        date_range_picker = DateRangePickerForm(request.POST)

        if 'date_range' in request.POST:
            if date_range_picker.is_valid():
                new_start = date_range_picker.cleaned_data['date_start']
                new_end = date_range_picker.cleaned_data['date_end']
                if driver_id == None:
                    return HttpResponseRedirect(reverse('report', kwargs={'start_year':new_start.year, 'start_month':new_start.month, 'start_day':new_start.day, 'end_year':new_end.year, 'end_month':new_end.month, 'end_day':new_end.day}))
                else:
                    return HttpResponseRedirect(reverse('report-driver', kwargs={'driver_id': driver_id, 'start_year':new_start.year, 'start_month':new_start.month, 'start_day':new_start.day, 'end_year':new_end.year, 'end_month':new_end.month, 'end_day':new_end.day}))
        elif 'show_daily_data' in request.POST:
            show_daily_data = request.session['report_show_daily_data'] = not request.session['report_show_daily_data']
            if driver_id == None:
                return HttpResponseRedirect(reverse('report', kwargs={'start_year':start_year, 'start_month':start_month, 'start_day':start_day, 'end_year':end_year, 'end_month':end_month, 'end_day':end_day}))
            else:
                return HttpResponseRedirect(reverse('report-driver', kwargs={'driver_id': driver_id, 'start_year':start_year, 'start_month':start_month, 'start_day':start_day, 'end_year':end_year, 'end_month':end_month, 'end_day':end_day}))
        else:
            if date_picker.is_valid():
                date_picker_date = date_picker.cleaned_data['date']
                if driver_id == None:
                    return HttpResponseRedirect(reverse('report-month', kwargs={'year':date_picker_date.year, 'month':date_picker_date.month}))
                else:
                    return HttpResponseRedirect(reverse('report-month-driver', kwargs={'driver_id': driver_id, 'year':date_picker_date.year, 'month':date_picker_date.month}))
    else:
        date_picker = DatePickerForm(initial={'date':date_start})
        date_range_picker = DateRangePickerForm(initial={'date_start':date_start, 'date_end':date_end})

    month_prev = date_start + datetime.timedelta(days=-1)
    month_prev.replace(day=1)
    month_next = date_end + datetime.timedelta(days=1)

    report = Report()
    report.load(date_start, date_end, driver_id=driver_id)

    if driver_id == None:
        url_month_prev = reverse('report-month', kwargs={'year': month_prev.year, 'month': month_prev.month})
        url_month_next = reverse('report-month', kwargs={'year': month_next.year, 'month': month_next.month})
        url_this_month = reverse('report-this-month')
        url_print = reverse('report-print', kwargs={'start_year': date_start.year, 'start_month': date_start.month, 'start_day': date_start.day, 'end_year': date_end.year, 'end_month': date_end.month, 'end_day': date_end.day})
        url_xlsx = reverse('report-xlsx', kwargs={'start_year': date_start.year, 'start_month': date_start.month, 'start_day': date_start.day, 'end_year': date_end.year, 'end_month': date_end.month, 'end_day': date_end.day})
    else:
        url_month_prev = reverse('report-month-driver', kwargs={'driver_id': driver_id, 'year': month_prev.year, 'month': month_prev.month})
        url_month_next = reverse('report-month-driver', kwargs={'driver_id': driver_id, 'year': month_next.year, 'month': month_next.month})
        url_this_month = reverse('report-this-month-driver', kwargs={'driver_id': driver_id})
        url_print = reverse('report-print-driver', kwargs={'driver_id': driver_id, 'start_year': date_start.year, 'start_month': date_start.month, 'start_day': date_start.day, 'end_year': date_end.year, 'end_month': date_end.month, 'end_day': date_end.day})
        url_xlsx = reverse('report-xlsx-driver', kwargs={'driver_id': driver_id, 'start_year': date_start.year, 'start_month': date_start.month, 'start_day': date_start.day, 'end_year': date_end.year, 'end_month': date_end.month, 'end_day': date_end.day})

    # we don't care about the driver for the mileage summary
    url_print_mile_summary = reverse('report-print-mileage-summary', kwargs={'start_year': date_start.year, 'start_month': date_start.month, 'start_day': date_start.day, 'end_year': date_end.year, 'end_month': date_end.month, 'end_day': date_end.day})

    selected_driver = None
    if driver_id != None:
        selected_driver = Driver.objects.get(id=driver_id)

    context = {
        'date_start': date_start,
        'date_end': date_end,
        'report': report,
        'date_picker': date_picker,
        'date_range_picker': date_range_picker,
        'url_month_prev': url_month_prev,
        'url_month_next': url_month_next,
        'url_this_month': url_this_month,
        'url_print': url_print,
        'url_print_mile_summary': url_print_mile_summary,
        'url_xlsx': url_xlsx,
        'selected_driver': selected_driver,
        'show_daily_data': show_daily_data,
        'is_short_report': len(report.report_all) <= 31,
        'vehicle_table_colspan': 13 + (3 * len(Report.ReportSummary.query_triptypes)),
        'drivers': Driver.objects.filter(is_active=True),
        'vehicles': Vehicle.objects.all(),
    }
    return render(request, 'report/view.html', context)



@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportPrint(request, start_year, start_month, start_day, end_year, end_month, end_day):
    return reportPrintBase(request, None, start_year, start_month, start_day, end_year, end_month, end_day)

@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportPrintDriver(request, driver_id, start_year, start_month, start_day, end_year, end_month, end_day):
    return reportPrintBase(request, driver_id, start_year, start_month, start_day, end_year, end_month, end_day)

@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportPrintBase(request, driver_id, start_year, start_month, start_day, end_year, end_month, end_day):
    date_start = datetime.date(start_year, start_month, start_day)
    date_end = datetime.date(end_year, end_month, end_day)

    show_daily_data = request.session.get('report_show_daily_data', False)
    request.session['report_show_daily_data'] = show_daily_data

    if date_start > date_end:
        swap_date = date_start
        date_start = date_end
        date_end = swap_date

    report = Report()
    report.load(date_start, date_end, driver_id=driver_id)

    context = {
        'date_start': date_start,
        'date_end': date_end,
        'report': report,
        'show_daily_data': show_daily_data,
        'is_short_report': len(report.report_all) <= 31,
        'vehicle_table_colspan': 13 + (3 * len(Report.ReportSummary.query_triptypes)),
        'vehicles': Vehicle.objects.all(),
    }
    return render(request, 'report/print.html', context)

@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportPrintMileageSummary(request, start_year, start_month, start_day, end_year, end_month, end_day):
    date_start = datetime.date(start_year, start_month, start_day)
    date_end = datetime.date(end_year, end_month, end_day)

    if date_start > date_end:
        swap_date = date_start
        date_start = date_end
        date_end = swap_date

    report = Report()
    report.load(date_start, date_end, driver_id=None)

    context = {
        'date_start': date_start,
        'date_end': date_end,
        'report': report,
    }
    return render(request, 'report/print_mileage_summary.html', context)



@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportMonth(request, year, month):
    return reportMonthBase(request, None, year, month)

@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportMonthDriver(request, driver_id, year, month):
    return reportMonthBase(request, driver_id, year, month)

@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportMonthBase(request, driver_id, year, month):
    date_start = datetime.date(year, month, 1)
    date_end = date_start
    if date_end.month == 12:
        date_end = date_end.replace(day=31)
    else:
        date_end = datetime.date(year, month+1, 1) + datetime.timedelta(days=-1)

    if driver_id:
        return HttpResponseRedirect(reverse('report-driver', kwargs={'driver_id': driver_id, 'start_year':date_start.year, 'start_month':date_start.month, 'start_day':date_start.day, 'end_year':date_end.year, 'end_month':date_end.month, 'end_day':date_end.day}))
    else:
        return HttpResponseRedirect(reverse('report', kwargs={'start_year':date_start.year, 'start_month':date_start.month, 'start_day':date_start.day, 'end_year':date_end.year, 'end_month':date_end.month, 'end_day':date_end.day}))


@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportYear(request, year):
    return reportYearBase(request, None, year)

@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportYearBase(request, driver_id, year):
    date_start = datetime.date(year, 1, 1)
    date_end = datetime.date(year+1, 1, 1) + datetime.timedelta(days=-1)

    if driver_id:
        return HttpResponseRedirect(reverse('report-driver', kwargs={'driver_id': driver_id, 'start_year':date_start.year, 'start_month':date_start.month, 'start_day':date_start.day, 'end_year':date_end.year, 'end_month':date_end.month, 'end_day':date_end.day}))
    else:
        return HttpResponseRedirect(reverse('report', kwargs={'start_year':date_start.year, 'start_month':date_start.month, 'start_day':date_start.day, 'end_year':date_end.year, 'end_month':date_end.month, 'end_day':date_end.day}))


@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportThisMonth(request):
    return reportThisMonthBase(request, None)

@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportThisMonthDriver(request, driver_id):
    return reportThisMonthBase(request, driver_id)

@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportThisMonthBase(request, driver_id):
    date = datetime.datetime.now().date()
    if driver_id:
        return reportMonthDriver(request, driver_id, date.year, date.month)
    else:
        return reportMonth(request, date.year, date.month)



@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportLastMonth(request):
    return reportLastMonthBase(request, None)

@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportLastMonthDriver(request, driver_id):
    return reportLastMonthBase(request, driver_id)

@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportLastMonthBase(request, driver_id):
    date = (datetime.datetime.now().date()).replace(day=1) # first day of this month 
    date = date + datetime.timedelta(days=-1) # last day of the previous month
    if driver_id:
        return reportMonthDriver(request, driver_id, date.year, date.month)
    else:
        return reportMonth(request, date.year, date.month)



@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportXLSX(request, start_year, start_month, start_day, end_year, end_month, end_day):
    return reportXLSXBase(request, None, start_year, start_month, start_day, end_year, end_month, end_day)

@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportXLSXDriver(request, driver_id, start_year, start_month, start_day, end_year, end_month, end_day):
    return reportXLSXBase(request, driver_id, start_year, start_month, start_day, end_year, end_month, end_day)

@permission_required(['transit.view_trip', 'transit.view_shift'])
def reportXLSXBase(request, driver_id, start_year, start_month, start_day, end_year, end_month, end_day):
    date_start = datetime.date(start_year, start_month, start_day)
    date_end = datetime.date(end_year, end_month, end_day)

    if date_start > date_end:
        swap_date = date_start
        date_start = date_end
        date_end = swap_date

    trip_types = TripType.objects.filter(is_trip_counted=True)
    tags = Tag.objects.all()

    report = Report()
    report.load(date_start, date_end, driver_id=driver_id)

    temp_file = tempfile.NamedTemporaryFile()

    wb = Workbook()

    style_font_normal = Font(name='Arial', size=10)
    style_border_normal_side = Side(border_style='thin', color='FF000000')
    style_border_normal = Border(left=style_border_normal_side, right=style_border_normal_side, top=style_border_normal_side, bottom=style_border_normal_side)
    style_colwidth_normal = 13
    style_colwidth_small = style_colwidth_normal / 3

    style_font_header = Font(name='Arial', size=9, bold=True)
    style_alignment_header = Alignment(horizontal='center', vertical='center', wrap_text=True)
    style_fill_header = PatternFill(fill_type='solid', fgColor='DFE0E1')
    style_rowheight_header = 25

    style_font_total = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
    style_fill_total = PatternFill(fill_type='solid', fgColor='27A343')

    #####
    #### All Vehicle Totals
    #####
    report_all = []
    for i in report.report_all:
        if i.hasVehicleInShift():
            report_all.append(i)

    ws = wb.active
    ws.title = 'Totals for All Vehicles'

    row_header = 1
    row_total = row_header + len(report_all) + 1

    trip_type_sub_cols = 3
    trip_type_cols = len(trip_types) * trip_type_sub_cols
    trip_type_start = 8
    trip_type_end = trip_type_start + trip_type_cols + trip_type_sub_cols

    col_count = 10 + trip_type_cols + trip_type_sub_cols

    ws.cell(row_header, 2, 'Service Miles')
    ws.cell(row_header, 3, 'Service Hours')
    ws.cell(row_header, 4, 'Deadhead Miles')
    ws.cell(row_header, 5, 'Deadhead Hours')
    ws.cell(row_header, 6, 'Passenger Miles (PMT)')
    ws.cell(row_header, 7, 'Fuel')
    ws.cell(row_header, trip_type_start + trip_type_cols, 'Total Trips')
    ws.cell(row_header, trip_type_end, 'Cash Collected')
    ws.cell(row_header, trip_type_end + 1, 'Check Collected')
    ws.cell(row_header, trip_type_end + 2, 'Total Money Collected')

    ws.row_dimensions[row_header].height = style_rowheight_header

    total_data = report.all_vehicles

    for i in range(0, row_total):
        row = i + 1

        # apply styles
        for col in range(1, col_count+1):
            ws.cell(row, col).border = style_border_normal

            if row == row_header:
                if col >= trip_type_start and col < trip_type_end:
                    ws.column_dimensions[get_column_letter(col)].width = style_colwidth_small

                    # merge trip type column headers
                    if (col-trip_type_start) % trip_type_sub_cols == 0:
                        ws.merge_cells(start_row=row_header, start_column=col, end_row=row_header, end_column=col+trip_type_sub_cols-1)
                else:
                    ws.column_dimensions[get_column_letter(col)].width = style_colwidth_normal

                ws.cell(row, col).font = style_font_header
                ws.cell(row, col).alignment = style_alignment_header
                ws.cell(row, col).fill = style_fill_header
            elif row == row_total:
                ws.cell(row, col).font = style_font_total
                ws.cell(row, col).fill = style_fill_total
            else:
                ws.cell(row, col).font = style_font_normal

            # number formats
            if row > row_header:
                if row < row_total:
                    ws.cell(row, 1).number_format = 'mmm dd, yyyy'
                ws.cell(row, 2).number_format = '0.0'
                ws.cell(row, 3).number_format = '0.00'
                ws.cell(row, 4).number_format = '0.0'
                ws.cell(row, 5).number_format = '0.00'
                ws.cell(row, 6).number_format = '0.0'
                ws.cell(row, 7).number_format = '0.0'
                ws.cell(row, trip_type_end).number_format = '$0.00'
                ws.cell(row, trip_type_end + 1).number_format = '$0.00'
                ws.cell(row, trip_type_end + 2).number_format = '$0.00'

        if row == row_header:
            continue

        if row == row_total:
            ws.cell(row_total, 1, 'TOTAL')
            ws.cell(row_total, 2, total_data.service_miles)
            ws.cell(row_total, 3, total_data.service_hours)
            ws.cell(row_total, 4, total_data.deadhead_miles)
            ws.cell(row_total, 5, total_data.deadhead_hours)
            ws.cell(row_total, 6, total_data.pmt)
            ws.cell(row_total, 7, total_data.fuel)

            for trip_type_index in range(0, len(trip_types)):
                trip_type = trip_types[trip_type_index]
                trip_type_col = trip_type_start + (trip_type_index * trip_type_sub_cols)
                ws.cell(row_header, trip_type_col, 'Trip Type: ' + str(trip_type))
                ws.cell(row_total, trip_type_col, total_data.trip_types[trip_type].passenger)
                ws.cell(row_total, trip_type_col + 1, total_data.trip_types[trip_type].no_passenger)
                ws.cell(row_total, trip_type_col + 2, total_data.trip_types[trip_type].total)

            ws.cell(row_total, trip_type_start + trip_type_cols, total_data.trip_types_total.passenger)
            ws.cell(row_total, trip_type_start + trip_type_cols + 1, total_data.trip_types_total.no_passenger)
            ws.cell(row_total, trip_type_start + trip_type_cols + 2, total_data.trip_types_total.total)

            ws.cell(row_total, trip_type_end, total_data.collected_cash.to_float())
            ws.cell(row_total, trip_type_end + 1, total_data.collected_check.to_float())
            ws.cell(row_total, trip_type_end + 2, total_data.total_collected_money.to_float())

            continue

        # daily report data
        rdata = report_all[i-1]
        rdata_date = rdata.date
        rdata_all = rdata.all

        ws.cell(row, 1, rdata_date)
        ws.cell(row, 2, rdata_all.service_miles)
        ws.cell(row, 3, rdata_all.service_hours)
        ws.cell(row, 4, rdata_all.deadhead_miles)
        ws.cell(row, 5, rdata_all.deadhead_hours)
        ws.cell(row, 6, rdata_all.pmt)
        ws.cell(row, 7, rdata_all.fuel)

        for trip_type_index in range(0, len(trip_types)):
            trip_type = trip_types[trip_type_index]
            trip_type_col = trip_type_start + (trip_type_index * trip_type_sub_cols)
            ws.cell(row, trip_type_col, rdata_all.trip_types[trip_type].passenger)
            ws.cell(row, trip_type_col + 1, rdata_all.trip_types[trip_type].no_passenger)
            ws.cell(row, trip_type_col + 2, rdata_all.trip_types[trip_type].total)

        ws.cell(row, trip_type_start + trip_type_cols, rdata_all.trip_types_total.passenger)
        ws.cell(row, trip_type_start + trip_type_cols + 1, rdata_all.trip_types_total.no_passenger)
        ws.cell(row, trip_type_start + trip_type_cols + 2, rdata_all.trip_types_total.total)

        ws.cell(row, trip_type_end, rdata_all.collected_cash.to_float())
        ws.cell(row, trip_type_end + 1, rdata_all.collected_check.to_float())
        ws.cell(row, trip_type_end + 2, rdata_all.total_collected_money.to_float())

    #####
    #### Per-Vehicle Reports
    #####
    for vr in report.vehicle_reports:
        if len(vr.days) == 0:
            continue

        ws = wb.create_sheet('Vehicle - ' + str(vr.vehicle))

        row_header = 1
        row_total = row_header + len(vr.days) + 1

        trip_type_sub_cols = 3
        trip_type_cols = len(trip_types) * trip_type_sub_cols
        trip_type_start = 8
        trip_type_end = trip_type_start + trip_type_cols + trip_type_sub_cols

        col_count = 10 + trip_type_cols + trip_type_sub_cols

        ws.cell(row_header, 2, 'Service Miles')
        ws.cell(row_header, 3, 'Service Hours')
        ws.cell(row_header, 4, 'Deadhead Miles')
        ws.cell(row_header, 5, 'Deadhead Hours')
        ws.cell(row_header, 6, 'Passenger Miles (PMT)')
        ws.cell(row_header, 7, 'Fuel')
        ws.cell(row_header, trip_type_start + trip_type_cols, 'Total Trips')
        ws.cell(row_header, trip_type_end, 'Cash Collected')
        ws.cell(row_header, trip_type_end + 1, 'Check Collected')
        ws.cell(row_header, trip_type_end + 2, 'Total Money Collected')

        ws.row_dimensions[row_header].height = style_rowheight_header

        total_data = vr.totals

        for i in range(0, row_total):
            row = i + 1

            # apply styles
            for col in range(1, col_count+1):
                ws.cell(row, col).border = style_border_normal

                if row == row_header:
                    if col >= trip_type_start and col < trip_type_end:
                        ws.column_dimensions[get_column_letter(col)].width = style_colwidth_small

                        # merge trip type column headers
                        if (col-trip_type_start) % trip_type_sub_cols == 0:
                            ws.merge_cells(start_row=row_header, start_column=col, end_row=row_header, end_column=col+trip_type_sub_cols-1)
                    else:
                        ws.column_dimensions[get_column_letter(col)].width = style_colwidth_normal

                    ws.cell(row, col).font = style_font_header
                    ws.cell(row, col).alignment = style_alignment_header
                    ws.cell(row, col).fill = style_fill_header
                elif row == row_total:
                    ws.cell(row, col).font = style_font_total
                    ws.cell(row, col).fill = style_fill_total
                else:
                    ws.cell(row, col).font = style_font_normal

                # number formats
                if row > row_header:
                    if row < row_total:
                        ws.cell(row, 1).number_format = 'mmm dd, yyyy'
                    ws.cell(row, 2).number_format = '0.0'
                    ws.cell(row, 3).number_format = '0.00'
                    ws.cell(row, 4).number_format = '0.0'
                    ws.cell(row, 5).number_format = '0.00'
                    ws.cell(row, 6).number_format = '0.0'
                    ws.cell(row, 7).number_format = '0.0'
                    ws.cell(row, trip_type_end).number_format = '$0.00'
                    ws.cell(row, trip_type_end + 1).number_format = '$0.00'
                    ws.cell(row, trip_type_end + 2).number_format = '$0.00'

            if row == row_header:
                continue

            if row == row_total:
                ws.cell(row_total, 1, 'TOTAL')
                ws.cell(row_total, 2, total_data.service_miles)
                ws.cell(row_total, 3, total_data.service_hours)
                ws.cell(row_total, 4, total_data.deadhead_miles)
                ws.cell(row_total, 5, total_data.deadhead_hours)
                ws.cell(row_total, 6, total_data.pmt)
                ws.cell(row_total, 7, total_data.fuel)

                for trip_type_index in range(0, len(trip_types)):
                    trip_type = trip_types[trip_type_index]
                    trip_type_col = trip_type_start + (trip_type_index * trip_type_sub_cols)
                    ws.cell(row_header, trip_type_col, 'Trip Type: ' + str(trip_type))
                    ws.cell(row_total, trip_type_col, total_data.trip_types[trip_type].passenger)
                    ws.cell(row_total, trip_type_col + 1, total_data.trip_types[trip_type].no_passenger)
                    ws.cell(row_total, trip_type_col + 2, total_data.trip_types[trip_type].total)

                ws.cell(row_total, trip_type_start + trip_type_cols, total_data.trip_types_total.passenger)
                ws.cell(row_total, trip_type_start + trip_type_cols + 1, total_data.trip_types_total.no_passenger)
                ws.cell(row_total, trip_type_start + trip_type_cols + 2, total_data.trip_types_total.total)

                ws.cell(row_total, trip_type_end, total_data.collected_cash.to_float())
                ws.cell(row_total, trip_type_end + 1, total_data.collected_check.to_float())
                ws.cell(row_total, trip_type_end + 2, total_data.total_collected_money.to_float())

                continue

            # daily report data
            rdata = vr.days[i-1]
            rdata_date = rdata['date']
            rdata_all = rdata['data']

            ws.cell(row, 1, rdata_date)
            ws.cell(row, 2, rdata_all.service_miles)
            ws.cell(row, 3, rdata_all.service_hours)
            ws.cell(row, 4, rdata_all.deadhead_miles)
            ws.cell(row, 5, rdata_all.deadhead_hours)
            ws.cell(row, 6, rdata_all.pmt)
            ws.cell(row, 7, rdata_all.fuel)

            for trip_type_index in range(0, len(trip_types)):
                trip_type = trip_types[trip_type_index]
                trip_type_col = trip_type_start + (trip_type_index * trip_type_sub_cols)
                ws.cell(row, trip_type_col, rdata_all.trip_types[trip_type].passenger)
                ws.cell(row, trip_type_col + 1, rdata_all.trip_types[trip_type].no_passenger)
                ws.cell(row, trip_type_col + 2, rdata_all.trip_types[trip_type].total)

            ws.cell(row, trip_type_start + trip_type_cols, rdata_all.trip_types_total.passenger)
            ws.cell(row, trip_type_start + trip_type_cols + 1, rdata_all.trip_types_total.no_passenger)
            ws.cell(row, trip_type_start + trip_type_cols + 2, rdata_all.trip_types_total.total)

            ws.cell(row, trip_type_end, rdata_all.collected_cash.to_float())
            ws.cell(row, trip_type_end + 1, rdata_all.collected_check.to_float())
            ws.cell(row, trip_type_end + 2, rdata_all.total_collected_money.to_float())

    #####
    #### Unique rider summary
    #####
    ws = wb.create_sheet('Rider Summary')

    row_header = 1
    row_total = row_header + 3

    row_header_trips = row_total + 2
    row_total_trips = row_header_trips + 3

    row_header_riders = row_total_trips + 2
    row_riders = row_header_riders + 1
    row_riders_end = row_riders + len(report.unique_riders.names)

    ws.row_dimensions[row_header].height = style_rowheight_header
    ws.row_dimensions[row_header_trips].height = style_rowheight_header
    ws.row_dimensions[row_header_riders].height = style_rowheight_header

    ws.cell(row_header, 2, 'Elderly Ambulatory')
    ws.cell(row_header, 3, 'Elderly Non-Ambulatory')
    ws.cell(row_header, 4, 'Non-Elderly Ambulatory')
    ws.cell(row_header, 5, 'Non-Elderly Non-Ambulatory')
    ws.cell(row_header, 6, 'Unknown')
    ws.cell(row_header, 7, 'Total')
    ws.cell(row_header, 8, 'Staff')
    ws.cell(row_header, 9, 'Total (with staff)')

    ws.cell(row_header+1, 1, 'On vehicle')
    ws.cell(row_header+2, 1, 'Not on vehicle')
    ws.cell(row_total, 1, 'TOTAL')
    for i in range(0, 8):
        ws.cell(row_header+1, i+2, report.unique_riders.by_individuals[i].passenger)
        ws.cell(row_header+2, i+2, report.unique_riders.by_individuals[i].no_passenger)
        ws.cell(row_total, i+2, report.unique_riders.by_individuals[i].total)

    ws.cell(row_header_trips, 2, 'Elderly Ambulatory')
    ws.cell(row_header_trips, 3, 'Elderly Non-Ambulatory')
    ws.cell(row_header_trips, 4, 'Non-Elderly Ambulatory')
    ws.cell(row_header_trips, 5, 'Non-Elderly Non-Ambulatory')
    ws.cell(row_header_trips, 6, 'Unknown')
    ws.cell(row_header_trips, 7, 'Total')
    ws.cell(row_header_trips, 8, 'Staff')
    ws.cell(row_header_trips, 9, 'Total (with staff)')

    ws.cell(row_header_trips+1, 1, 'Trips on vehicle')
    ws.cell(row_header_trips+2, 1, 'Not on vehicle')
    ws.cell(row_total, 1, 'TOTAL')
    for i in range(0,8):
        ws.cell(row_header_trips+1, i+2, report.unique_riders.by_trips[i].passenger)
        ws.cell(row_header_trips+2, i+2, report.unique_riders.by_trips[i].no_passenger)
        ws.cell(row_total_trips, i+2, report.unique_riders.by_trips[i].total)

    ws.cell(row_header_riders, 1, 'Name')
    ws.cell(row_header_riders, 3, 'Elderly')
    ws.cell(row_header_riders, 4, 'Ambulatory')
    ws.cell(row_header_riders, 5, 'Trips on vehicle')
    ws.cell(row_header_riders, 6, 'Trips not on vehicle')
    ws.cell(row_header_riders, 7, 'No-Show')
    ws.cell(row_header_riders, 8, 'Canceled (late)')
    ws.cell(row_header_riders, 9, 'Canceled (same-day)')
    ws.cell(row_header_riders, 10, 'Total Trips')

    rider_index = 0

    unique_rider_list = list(report.unique_riders.names.values())

    for i in range(0, row_riders_end):
        row = i + 1

        # apply styles
        if row <= row_total_trips:
            for col in range(1, 10):
                ws.cell(row, col).border = style_border_normal
                if row == row_header or row == row_header_trips:
                    ws.column_dimensions[get_column_letter(col)].width = style_colwidth_normal
                    ws.cell(row, col).font = style_font_header
                    ws.cell(row, col).alignment = style_alignment_header
                    ws.cell(row, col).fill = style_fill_header
                elif row == row_total or row == row_total_trips:
                    ws.cell(row, col).font = style_font_total
                    ws.cell(row, col).fill = style_fill_total
                else:
                    ws.cell(row, col).font = style_font_normal
            continue
        elif row == row_header_riders:
            for col in range(1, 11):
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                ws.cell(row, col).border = style_border_normal
                ws.cell(row, col).font = style_font_header
                ws.cell(row, col).alignment = style_alignment_header
                ws.cell(row, col).fill = style_fill_header
            continue
        elif row < row_header_riders:
            continue

        if rider_index >= len(report.unique_riders.names):
            break

        while rider_index < len(unique_rider_list):
            rdata = unique_rider_list[rider_index]
            rider_index += 1
            if rdata.trips.total > 0:
                ws.cell(row, 1, rdata.name)
                ws.cell(row, 3, rdata.elderly)
                ws.cell(row, 4, rdata.ambulatory)
                ws.cell(row, 5, rdata.trips.passenger)
                ws.cell(row, 6, rdata.trips.no_passenger)
                ws.cell(row, 7, rdata.trips_no_show.total)
                ws.cell(row, 8, rdata.trips_canceled_late.total)
                ws.cell(row, 9, rdata.trips_canceled_very_late.total)
                ws.cell(row, 10, rdata.trips.total)

                # apply rider row styles
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                for col in range(1, 11):
                    ws.cell(row, col).border = style_border_normal
                    ws.cell(row, col).font = style_font_normal
                    if col == 3 or col == 4:
                        ws.cell(row, col).number_format = '0'

                break

    #####
    #### Trip Types and Tags
    #####
    ws = wb.create_sheet('Trip Types and Tags')

    row_header = 1
    row_trip_type = row_header + 1
    row_trip_type_unknown = row_trip_type + len(trip_types)
    row_trip_type_total = row_trip_type_unknown + 1
    row_header_tags = row_trip_type_total + 2
    row_tags = row_header_tags + 1
    row_tags_end = row_tags + len(tags)

    ws.cell(row_header, 1, 'Trip Type')
    ws.cell(row_header, 2, 'Trips on vehicle')
    ws.cell(row_header, 3, 'Trips not on vehicle')
    ws.cell(row_header, 4, 'Total Trips')

    ws.cell(row_trip_type_unknown, 1, 'Unknown')
    ws.cell(row_trip_type_unknown, 2, report.all_vehicles.trip_types_unknown.passenger)
    ws.cell(row_trip_type_unknown, 3, report.all_vehicles.trip_types_unknown.no_passenger)
    ws.cell(row_trip_type_unknown, 4, report.all_vehicles.trip_types_unknown.total)

    ws.cell(row_trip_type_total, 1, 'TOTAL')
    ws.cell(row_trip_type_total, 2, report.all_vehicles.trip_types_total.passenger)
    ws.cell(row_trip_type_total, 3, report.all_vehicles.trip_types_total.no_passenger)
    ws.cell(row_trip_type_total, 4, report.all_vehicles.trip_types_total.total)

    ws.cell(row_header_tags, 1, 'Tag')
    ws.cell(row_header_tags, 2, 'Trips on vehicle')
    ws.cell(row_header_tags, 3, 'Trips not on vehicle')
    ws.cell(row_header_tags, 4, 'Total Trips')

    ws.row_dimensions[row_header].height = style_rowheight_header
    ws.row_dimensions[row_header_tags].height = style_rowheight_header
    ws.column_dimensions[get_column_letter(1)].width = style_colwidth_normal * 2

    for i in range(0, row_tags_end):
        row = i + 1

        if row == row_trip_type_total + 1:
            continue

        # apply styles
        for col in range(1, 5):

            if col > 1:
                ws.column_dimensions[get_column_letter(col)].width = style_colwidth_normal

            if row == row_trip_type_total:
                ws.cell(row, col).font = style_font_total
                ws.cell(row, col).fill = style_fill_total
                ws.cell(row, col).border = style_border_normal
            elif row == row_header or row == row_header_tags:
                ws.cell(row, col).font = style_font_header
                ws.cell(row, col).alignment = style_alignment_header
                ws.cell(row, col).fill = style_fill_header
            else:
                ws.cell(row, col).font = style_font_normal

            if row < row_tags_end:
                ws.cell(row, col).border = style_border_normal

        if row >= row_trip_type and row < row_trip_type_unknown:
            trip_type = trip_types[row - row_trip_type]
            ws.cell(row, 1, str(trip_type))
            ws.cell(row, 2, report.all_vehicles.trip_types[trip_type].passenger)
            ws.cell(row, 3, report.all_vehicles.trip_types[trip_type].no_passenger)
            ws.cell(row, 4, report.all_vehicles.trip_types[trip_type].total)
        elif row >= row_tags and row < row_tags_end:
            tag = str(tags[row - row_tags])
            ws.cell(row, 1, tag)
            ws.cell(row, 2, report.all_vehicles.tags[tag].passenger)
            ws.cell(row, 3, report.all_vehicles.tags[tag].no_passenger)
            ws.cell(row, 4, report.all_vehicles.tags[tag].total)

    #####
    #### Frequent Destinations
    #####
    ws = wb.create_sheet('Frequent Destinations')

    row_header = 1
    row_dest = row_header + 1
    row_dest_end = row_dest + len(report.frequent_destinations)

    ws.cell(row_header, 1, 'Address')
    ws.cell(row_header, 2, 'Trips on vehicle')
    ws.cell(row_header, 3, 'Trips not on vehicle')
    ws.cell(row_header, 4, 'Total Trips')
    ws.cell(row_header, 5, 'Average Mileage')

    ws.row_dimensions[row_header].height = style_rowheight_header
    ws.column_dimensions[get_column_letter(1)].width = style_colwidth_normal * 2

    frequent_destinations = list(report.frequent_destinations.values())
    for i in range(0, row_dest_end):
        row = i + 1

        # apply styles
        for col in range(1, 6):
            if col > 1:
                ws.column_dimensions[get_column_letter(col)].width = style_colwidth_normal

            if row < row_dest_end:
                ws.cell(row, col).border = style_border_normal

            if row == row_header:
                ws.cell(row, col).font = style_font_header
                ws.cell(row, col).alignment = style_alignment_header
                ws.cell(row, col).fill = style_fill_header
            else:
                ws.cell(row, col).font = style_font_normal

        if row < row_dest or row == row_dest_end:
            continue

        rdata = frequent_destinations[i-1]

        ws.cell(row, 1, rdata.address)
        ws.cell(row, 2, rdata.trips.passenger)
        ws.cell(row, 3, rdata.trips.no_passenger)
        ws.cell(row, 4, rdata.trips.total)
        ws.cell(row, 5, rdata.avg_mileage)
        ws.cell(row, 5).number_format = '0.0'

    #####
    #### Fares & Payments (by client)
    #####
    ws = wb.create_sheet('Fares & Payments (by client)')

    row_header = 1
    row_data = row_header + 1
    row_total = row_data
    row_data_max = row_data + len(report.unique_riders.names)

    ws.cell(row_header, 1, 'Name')
    ws.cell(row_header, 2, 'Cash (driver collected)')
    ws.cell(row_header, 3, 'Check (driver collected)')
    ws.cell(row_header, 4, 'Cash (not driver collected)')
    ws.cell(row_header, 5, 'Check (not driver collected)')
    ws.cell(row_header, 6, 'Total Payments')
    ws.cell(row_header, 7, 'Total Fares')
    ws.cell(row_header, 8, 'Total Owed')

    ws.row_dimensions[row_header].height = style_rowheight_header
    ws.column_dimensions[get_column_letter(1)].width = style_colwidth_normal * 2

    data_index = 0

    for i in range(0, row_data_max):
        row = i + 1

        if row == row_header:
            # apply styles
            for col in range(1, 9):
                if col > 1:
                    ws.column_dimensions[get_column_letter(col)].width = style_colwidth_normal
                ws.cell(row, col).border = style_border_normal
                ws.cell(row, col).font = style_font_header
                ws.cell(row, col).alignment = style_alignment_header
                ws.cell(row, col).fill = style_fill_header
        else:
            row_total = row

            while data_index < len(unique_rider_list):
                rdata = unique_rider_list[data_index]
                data_index += 1

                if rdata.total_payments.value > 0 or rdata.total_fares.value > 0:
                    ws.cell(row, 1, rdata.name)
                    ws.cell(row, 2, rdata.collected_cash.to_float())
                    ws.cell(row, 3, rdata.collected_check.to_float())
                    ws.cell(row, 4, rdata.paid_cash.to_float())
                    ws.cell(row, 5, rdata.paid_check.to_float())
                    ws.cell(row, 6, rdata.total_payments.to_float())
                    ws.cell(row, 7, rdata.total_fares.to_float())
                    ws.cell(row, 8, rdata.total_owed.to_float())

                    # apply styles
                    for col in range(1, 9):
                        if col > 1:
                            ws.cell(row, col).number_format = '$0.00'
                        ws.cell(row, col).border = style_border_normal
                        ws.cell(row, col).font = style_font_normal

                    row_total += 1

                    break

            if row == row_total:
                ws.cell(row, 1, 'TOTAL')
                ws.cell(row, 2, report.unique_riders.total_collected_cash.to_float())
                ws.cell(row, 3, report.unique_riders.total_collected_check.to_float())
                ws.cell(row, 4, report.unique_riders.total_paid_cash.to_float())
                ws.cell(row, 5, report.unique_riders.total_paid_check.to_float())
                ws.cell(row, 6, report.unique_riders.total_total_payments.to_float())
                ws.cell(row, 7, report.unique_riders.total_total_fares.to_float())
                ws.cell(row, 8, report.unique_riders.total_total_owed.to_float())

                # apply styles
                for col in range(1, 9):
                    if col > 1:
                        ws.cell(row, col).number_format = '$0.00'
                    ws.cell(row, col).border = style_border_normal
                    ws.cell(row, col).font = style_font_total
                    ws.cell(row, col).fill = style_fill_total

                break

    #####
    #### Fares & Payments (by date)
    #####
    ws = wb.create_sheet('Fares & Payments (by date)')

    row_header = 1
    row_data = row_header + 1
    row_total = row_data
    row_data_max = row_data + len(report.report_all)

    ws.cell(row_header, 1, 'Date')
    ws.cell(row_header, 2, 'Cash (driver collected)')
    ws.cell(row_header, 3, 'Check (driver collected)')
    ws.cell(row_header, 4, 'Cash (not driver collected)')
    ws.cell(row_header, 5, 'Check (not driver collected)')
    ws.cell(row_header, 6, 'Total Payments')
    ws.cell(row_header, 7, 'Total Fares')

    ws.row_dimensions[row_header].height = style_rowheight_header

    data_index = 0

    for i in range(0, row_data_max):
        row = i + 1

        if row == row_header:
            # apply styles
            for col in range(1, 8):
                ws.column_dimensions[get_column_letter(col)].width = style_colwidth_normal
                ws.cell(row, col).border = style_border_normal
                ws.cell(row, col).font = style_font_header
                ws.cell(row, col).alignment = style_alignment_header
                ws.cell(row, col).fill = style_fill_header
        else:
            row = row_total

            while data_index < len(report.report_all):
                rdata = report.report_all[data_index]
                data_index += 1

                if rdata.total_payments.value > 0 or rdata.total_fares.value > 0:
                    ws.cell(row, 1, rdata.date)
                    ws.cell(row, 2, rdata.collected_cash.to_float())
                    ws.cell(row, 3, rdata.collected_check.to_float())
                    ws.cell(row, 4, rdata.paid_cash.to_float())
                    ws.cell(row, 5, rdata.paid_check.to_float())
                    ws.cell(row, 6, rdata.total_payments.to_float())
                    ws.cell(row, 7, rdata.total_fares.to_float())

                    # apply styles
                    for col in range(1, 8):
                        if col == 1:
                            ws.cell(row, col).number_format = 'mmm dd, yyyy'
                        else:
                            ws.cell(row, col).number_format = '$0.00'
                        ws.cell(row, col).border = style_border_normal
                        ws.cell(row, col).font = style_font_normal

                    row_total += 1

                    break

            if row == row_total:
                ws.cell(row, 1, 'TOTAL')
                ws.cell(row, 2, report.unique_riders.total_collected_cash.to_float())
                ws.cell(row, 3, report.unique_riders.total_collected_check.to_float())
                ws.cell(row, 4, report.unique_riders.total_paid_cash.to_float())
                ws.cell(row, 5, report.unique_riders.total_paid_check.to_float())
                ws.cell(row, 6, report.unique_riders.total_total_payments.to_float())
                ws.cell(row, 7, report.unique_riders.total_total_fares.to_float())

                # apply styles
                for col in range(1, 8):
                    if col > 1:
                        ws.cell(row, col).number_format = '$0.00'
                    ws.cell(row, col).border = style_border_normal
                    ws.cell(row, col).font = style_font_total
                    ws.cell(row, col).fill = style_fill_total

                break

    #####
    #### Money Collected by the Drivers
    #####
    ws = wb.create_sheet('Driver-collected Money')

    row_header = 1
    row_data = row_header + 1
    row_total = row_data
    row_data_max = row_data + len(report.money_trips)

    ws.cell(row_header, 1, 'Date')
    ws.cell(row_header, 2, 'Name')
    ws.cell(row_header, 3, 'Cash')
    ws.cell(row_header, 4, 'Check')

    ws.row_dimensions[row_header].height = style_rowheight_header
    ws.column_dimensions[get_column_letter(2)].width = style_colwidth_normal * 2

    data_index = 0

    for i in range(0, row_data_max):
        row = i + 1

        if row == row_header:
            # apply styles
            for col in range(1, 5):
                if col != 2:
                    ws.column_dimensions[get_column_letter(col)].width = style_colwidth_normal
                ws.cell(row, col).border = style_border_normal
                ws.cell(row, col).font = style_font_header
                ws.cell(row, col).alignment = style_alignment_header
                ws.cell(row, col).fill = style_fill_header
        else:
            row_total = row

            if data_index < len(report.money_trips):
                rdata = report.money_trips[data_index]
                data_index += 1

                ws.cell(row, 1, rdata.trip.date)
                ws.cell(row, 2, rdata.trip.name)
                ws.cell(row, 3, rdata.collected_cash.to_float())
                ws.cell(row, 4, rdata.collected_check.to_float())

                # apply styles
                for col in range(1, 5):
                    if col == 1:
                        ws.cell(row, col).number_format = 'mmm dd, yyyy'
                    elif col == 3 or col == 4:
                        ws.cell(row, col).number_format = '$0.00'
                    ws.cell(row, col).border = style_border_normal
                    ws.cell(row, col).font = style_font_normal

                row_total += 1

            if row == row_total:
                ws.cell(row, 1, 'TOTAL')
                ws.cell(row, 3, report.money_trips_summary.collected_cash.to_float())
                ws.cell(row, 4, report.money_trips_summary.collected_check.to_float())

                # apply styles
                for col in range(1, 5):
                    if col == 3 or col == 4:
                        ws.cell(row, col).number_format = '$0.00'
                    ws.cell(row, col).border = style_border_normal
                    ws.cell(row, col).font = style_font_total
                    ws.cell(row, col).fill = style_fill_total

                break

    #####
    #### Money Not Collected by the Drivers
    #####
    ws = wb.create_sheet('Non-Driver-collected Money')

    row_header = 1
    row_data = row_header + 1
    row_total = row_data
    row_data_max = row_data + len(report.money_payments)

    ws.cell(row_header, 1, 'Date')
    ws.cell(row_header, 2, 'Name')
    ws.cell(row_header, 3, 'Cash')
    ws.cell(row_header, 4, 'Check')

    ws.row_dimensions[row_header].height = style_rowheight_header
    ws.column_dimensions[get_column_letter(2)].width = style_colwidth_normal * 2

    data_index = 0

    for i in range(0, row_data_max):
        row = i + 1

        if row == row_header:
            # apply styles
            for col in range(1, 5):
                if col != 2:
                    ws.column_dimensions[get_column_letter(col)].width = style_colwidth_normal
                ws.cell(row, col).border = style_border_normal
                ws.cell(row, col).font = style_font_header
                ws.cell(row, col).alignment = style_alignment_header
                ws.cell(row, col).fill = style_fill_header
        else:
            row_total = row

            if data_index < len(report.money_payments):
                rdata = report.money_payments[data_index]
                data_index += 1

                ws.cell(row, 1, rdata.date)
                ws.cell(row, 2, rdata.client.name)
                ws.cell(row, 3, rdata.cash.to_float())
                ws.cell(row, 4, rdata.check.to_float())

                # apply styles
                for col in range(1, 5):
                    if col == 1:
                        ws.cell(row, col).number_format = 'mmm dd, yyyy'
                    elif col == 3 or col == 4:
                        ws.cell(row, col).number_format = '$0.00'
                    ws.cell(row, col).border = style_border_normal
                    ws.cell(row, col).font = style_font_normal

                row_total += 1

            if row == row_total:
                ws.cell(row, 1, 'TOTAL')
                ws.cell(row, 3, report.money_payments_summary.cash.to_float())
                ws.cell(row, 4, report.money_payments_summary.check.to_float())

                # apply styles
                for col in range(1, 5):
                    if col == 3 or col == 4:
                        ws.cell(row, col).number_format = '$0.00'
                    ws.cell(row, col).border = style_border_normal
                    ws.cell(row, col).font = style_font_total
                    ws.cell(row, col).fill = style_fill_total

                break

    #####
    #### Per-Driver Summary
    #####
    ws = wb.create_sheet('Per-Driver Summary')

    row_header = 1
    row_data = row_header + 1
    row_total = row_data
    row_data_max = row_data + len(report.driver_reports)

    ws.cell(row_header, 1, 'Driver')
    ws.cell(row_header, 2, 'Service Miles')
    ws.cell(row_header, 3, 'Service Hours')
    ws.cell(row_header, 4, 'Deadhead Miles')
    ws.cell(row_header, 5, 'Deadhead Hours')
    ws.cell(row_header, 6, 'Total Miles')
    ws.cell(row_header, 7, 'Total Hours')
    ws.cell(row_header, 8, 'Days of Service')

    ws.row_dimensions[row_header].height = style_rowheight_header

    data_index = 0

    for i in range(0, row_data_max):
        row = i + 1

        if row == row_header:
            # apply styles
            for col in range(1, 9):
                ws.column_dimensions[get_column_letter(col)].width = style_colwidth_normal
                ws.cell(row, col).border = style_border_normal
                ws.cell(row, col).font = style_font_header
                ws.cell(row, col).alignment = style_alignment_header
                ws.cell(row, col).fill = style_fill_header
        else:
            row_total = row

            if data_index < len(report.driver_reports):
                rdata = report.driver_reports[data_index]
                data_index += 1

                ws.cell(row, 1, str(rdata.driver))
                ws.cell(row, 2, rdata.totals.service_miles)
                ws.cell(row, 3, rdata.totals.service_hours)
                ws.cell(row, 4, rdata.totals.deadhead_miles)
                ws.cell(row, 5, rdata.totals.deadhead_hours)
                ws.cell(row, 6, rdata.totals.total_miles)
                ws.cell(row, 7, rdata.totals.total_hours)
                ws.cell(row, 8, len(rdata.days))

                # apply styles
                style_fill_driver = PatternFill(fill_type='solid', fgColor=rdata.driver.get_color())
                for col in range(1, 9):
                    ws.cell(row, col).border = style_border_normal
                    ws.cell(row, col).font = style_font_normal
                    ws.cell(row, col).fill = style_fill_driver
                    if col == 2 or col == 4 or col == 6:
                        ws.cell(row, col).number_format = '0.0'
                    if col == 3 or col == 5 or col == 7:
                        ws.cell(row, col).number_format = '0.00'

                row_total += 1

            if row == row_total:
                ws.cell(row, 1, 'TOTAL')
                ws.cell(row, 2, report.driver_reports_total.totals.service_miles)
                ws.cell(row, 3, report.driver_reports_total.totals.service_hours)
                ws.cell(row, 4, report.driver_reports_total.totals.deadhead_miles)
                ws.cell(row, 5, report.driver_reports_total.totals.deadhead_hours)
                ws.cell(row, 6, report.driver_reports_total.totals.total_miles)
                ws.cell(row, 7, report.driver_reports_total.totals.total_hours)
                ws.cell(row, 8, report.total_vehicle_days_of_service)

                # apply styles
                for col in range(1, 9):
                    ws.cell(row, col).border = style_border_normal
                    ws.cell(row, col).font = style_font_total
                    ws.cell(row, col).fill = style_fill_total
                    if col == 2 or col == 4 or col == 6:
                        ws.cell(row, col).number_format = '0.0'
                    if col == 3 or col == 5 or col == 7:
                        ws.cell(row, col).number_format = '0.00'

                break

    #####
    #### Trips per Weekday
    #####
    ws = wb.create_sheet('Trips per Weekday')

    row_header = 1
    row_data = row_header + 1
    row_total = row_data

    ws.cell(row_header, 1, 'Monday')
    ws.cell(row_header, 2, 'Tuesday')
    ws.cell(row_header, 3, 'Wednesday')
    ws.cell(row_header, 4, 'Thursday')
    ws.cell(row_header, 5, 'Friday')
    ws.cell(row_header, 6, 'Saturday')
    ws.cell(row_header, 7, 'Sunday')

    ws.row_dimensions[row_header].height = style_rowheight_header

    data_index = 0

    for i in range(0, row_data_max):
        row = i + 1

        if row == row_header:
            # apply styles
            for col in range(1, 8):
                ws.column_dimensions[get_column_letter(col)].width = style_colwidth_normal
                ws.cell(row, col).border = style_border_normal
                ws.cell(row, col).font = style_font_header
                ws.cell(row, col).alignment = style_alignment_header
                ws.cell(row, col).fill = style_fill_header
        else:
            row_total = row

            if row == row_total:
                for col in range(1, 8):
                    ws.cell(row, col, report.weekday_totals[col-1].total)

                    # apply styles
                    ws.cell(row, col).border = style_border_normal
                    ws.cell(row, col).font = style_font_normal

                break

    wb.save(filename=temp_file.name)

    return FileResponse(open(temp_file.name, 'rb'), filename='Transit_Report_' + date_start.strftime('%Y-%m-%d') + '_to_' + date_end.strftime('%Y-%m-%d') + '.xlsx', as_attachment=True)
